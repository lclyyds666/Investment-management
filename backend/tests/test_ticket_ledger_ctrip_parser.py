import asyncio
import unittest
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from types import SimpleNamespace
from unittest.mock import patch

from fastapi import UploadFile
from openpyxl import Workbook

from app.api.v1.endpoints import ticket_ledger as ticket_endpoint
from app.services import ticket_ledger


class TicketLedgerCtripParserTest(unittest.TestCase):
    @staticmethod
    def _workbook_bytes() -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "抖音明细7.3-7.12"
        ws.append(["订单实收金额", "软件服务费", "达人服务费", "团长服务费", "核销时间"])
        ws.append([100, -1, -2, -3, datetime(2026, 7, 3, 10, 0)])

        headers = ["结算价金额", "流水类型", "服务完成日期", "出发时间", "付款日期"]
        ctrip_sheets = (
            ("携程明细7.3-7.12", [77] * 9 + [81], datetime(2026, 7, 3)),
            ("携程明细7.13-7.19", [62] * 15 + [75], datetime(2026, 7, 13)),
            ("携程明细7.20-7.26", [66] * 16 + [68], datetime(2026, 7, 26)),
        )
        for title, amounts, departure in ctrip_sheets:
            sheet = wb.create_sheet(title)
            sheet.append(headers)
            for amount in amounts:
                # 真实文件的服务完成日期为空，应回退到出发时间。
                sheet.append([amount, "订单成本", None, departure, datetime(2026, 7, 1)])
        # 非订单成本流水不得进入门票核销金额。
        wb["携程明细7.20-7.26"].append(
            [999, "调账", None, datetime(2026, 7, 26), datetime(2026, 7, 1)]
        )

        output = BytesIO()
        wb.save(output)
        wb.close()
        return output.getvalue()

    def test_mixed_file_is_split_by_platform_and_ctrip_is_calculated(self):
        parsed = ticket_ledger.parse_reconciliation(
            self._workbook_bytes(), "对账明细-2026.07.03-2026.07.25.xlsx"
        )
        self.assertEqual([item["platform"] for item in parsed["platforms"]], ["抖音", "携程"])

        douyin = parsed["platforms"][0]
        self.assertEqual(douyin["supplier_received"], Decimal("94.00"))
        self.assertEqual(douyin["suggested_commission"], Decimal("1.00"))

        ctrip = parsed["platforms"][1]
        self.assertEqual(ctrip["supplier_received"], Decimal("2903.00"))
        self.assertEqual(ctrip["suggested_commission"], Decimal("0.00"))
        self.assertEqual(ctrip["order_count"], 43)
        self.assertEqual(ctrip["positive_count"], 43)
        self.assertEqual(ctrip["period_start"].isoformat(), "2026-07-03")
        # 台账期次仍以文件名为准，7月26日明细只参与金额计算。
        self.assertEqual(ctrip["period_end"].isoformat(), "2026-07-25")
        self.assertTrue(ctrip["daily_json"])

        calculated = ticket_ledger.recompute_from_json(
            ctrip["daily_json"],
            Decimal("0.90"),
            Decimal("0.94"),
            ctrip["suggested_commission"],
            platform="携程",
            scenic_id="test-scenic",
        )
        self.assertEqual(calculated["publisher_due"], Decimal("2903.00"))
        self.assertEqual(calculated["hexiao_amount"], Decimal("2612.70"))
        self.assertEqual(calculated["jinying_amount"], Decimal("2728.82"))
        self.assertEqual(calculated["service_fee"], Decimal("116.12"))

    def test_running_balance_groups_platform_rows_as_one_period(self):
        rows = [
            SimpleNamespace(source_file="period-1.xlsx", payment_amount=5000, hexiao_amount=100),
            SimpleNamespace(source_file="period-1.xlsx", payment_amount=5000, hexiao_amount=200),
            SimpleNamespace(source_file="period-2.xlsx", payment_amount=0, hexiao_amount=50),
        ]
        balances = ticket_ledger.calculate_running_balances(
            "test-scenic", rows, group_by=lambda row: row.source_file
        )
        self.assertEqual(
            balances,
            [Decimal("4700.00"), Decimal("4700.00"), Decimal("4650.00")],
        )

    def test_parse_api_returns_one_draft_per_platform(self):
        upload = UploadFile(
            filename="对账明细-2026.07.03-2026.07.25.xlsx",
            file=BytesIO(self._workbook_bytes()),
        )
        with TemporaryDirectory() as temp_dir:
            with patch.object(ticket_endpoint, "_detail_dir", return_value=Path(temp_dir)):
                response = asyncio.run(
                    ticket_endpoint.parse_files(
                        scenic_id="test-scenic", files=[upload], db=None, _=None
                    )
                )
        self.assertEqual(response.code, 0)
        self.assertEqual(response.data.succeeded, 2)
        self.assertEqual(
            [item.platform for item in response.data.files], ["抖音", "携程"]
        )
        self.assertEqual(response.data.files[1].supplier_received, Decimal("2903.00"))
        self.assertTrue(response.data.files[1].daily_json)

    def test_sheet_names_without_detail_and_four_platforms_are_supported(self):
        wb = Workbook()

        douyin = wb.active
        douyin.title = "抖音5.25-5.31"
        douyin.append(["订单实收金额", "软件服务费", "达人服务费", "团长服务费", "核销时间"])
        douyin.append([100, -1, -2, -3, datetime(2026, 5, 25, 10, 0)])

        meituan = wb.create_sheet("美团5.25-5.31")
        meituan.append(["遵义动物园美团结算明细"])
        meituan.append(["结算方式", "应付金额", "张数", "时间"])
        meituan.append(["消费结算", 74.26, 2, datetime(2026, 5, 26, 10, 0)])
        meituan.append(["退款结算", -74.26, 2, datetime(2026, 5, 26, 11, 0)])

        ctrip = wb.create_sheet("携程6.1-6.7")
        ctrip.append(["结算价金额", "流水类型", "使用份数", "服务完成日期", "出发时间", "付款日期"])
        ctrip.append([49, "订单成本", 3, None, datetime(2026, 6, 1), datetime(2026, 6, 5)])

        tongcheng = wb.create_sheet("同程6.1-6.15")
        tongcheng.append(["商家应收", "订单票数", "旅游日期"])
        tongcheng.append([92.12, 2, datetime(2026, 6, 10)])

        output = BytesIO()
        wb.save(output)
        wb.close()

        parsed = ticket_ledger.parse_reconciliation(
            output.getvalue(), "遵义动物园5.25-6.21.xlsx"
        )
        by_platform = {item["platform"]: item for item in parsed["platforms"]}

        self.assertEqual(list(by_platform), ["抖音", "美团", "携程", "同程"])
        self.assertEqual(by_platform["抖音"]["supplier_received"], Decimal("94.00"))
        self.assertEqual(by_platform["美团"]["supplier_received"], Decimal("74.26"))
        self.assertEqual(by_platform["美团"]["order_count"], 2)
        self.assertEqual(by_platform["携程"]["supplier_received"], Decimal("49.00"))
        self.assertEqual(by_platform["携程"]["order_count"], 3)
        self.assertEqual(by_platform["同程"]["supplier_received"], Decimal("92.12"))
        self.assertEqual(by_platform["同程"]["order_count"], 2)
        for item in by_platform.values():
            self.assertEqual(item["period_start"].isoformat(), "2026-05-25")
            self.assertEqual(item["period_end"].isoformat(), "2026-06-21")
            self.assertTrue(item["daily_json"])

    @staticmethod
    def _target_scenic_workbook(include_meituan_tech_fee: bool = True) -> bytes:
        wb = Workbook()
        douyin = wb.active
        douyin.title = "抖音"
        douyin.append([
            "订单实收金额", "软件服务费", "达人服务费", "团长服务费",
            "服务商服务费", "核销时间",
        ])
        douyin.append([100, -1, -2, -3, -5, datetime(2026, 8, 1, 10, 0)])

        meituan = wb.create_sheet("美团")
        headers = ["结算方式", "应付金额"]
        values = ["消费结算", 80]
        if include_meituan_tech_fee:
            headers.append("技术服务费")
            values.append(-3)
        headers.extend(["张数", "时间"])
        values.extend([1, datetime(2026, 8, 1, 11, 0)])
        meituan.append(headers)
        meituan.append(values)

        output = BytesIO()
        wb.save(output)
        wb.close()
        return output.getvalue()

    def test_zunyi_uses_configured_rates_and_fixed_platform_received_rules(self):
        parsed = ticket_ledger.parse_reconciliation(
            self._target_scenic_workbook(),
            "遵义动物园8.1-8.1.xlsx",
            scenic_id="zunyi-zoo",
            rate_hexiao=Decimal("0.84"),
            rate_settle=Decimal("0.87"),
            commission_rate=Decimal("0"),
            commission_override=Decimal("0"),
            ticket_product="遵义动物园",
        )
        by_platform = {item["platform"]: item for item in parsed["platforms"]}

        douyin = by_platform["抖音"]
        self.assertEqual(douyin["supplier_received"], Decimal("93.00"))
        self.assertEqual(douyin["suggested_commission"], Decimal("0.00"))
        self.assertEqual(douyin["def_hexiao"], Decimal("78.12"))
        self.assertEqual(douyin["def_jinying"], Decimal("80.91"))
        self.assertEqual(douyin["ticket_product"], "遵义动物园")
        self.assertEqual(douyin["rate_hexiao"], Decimal("0.84"))
        self.assertEqual(douyin["rate_settle"], Decimal("0.87"))

        meituan = by_platform["美团"]
        self.assertEqual(meituan["supplier_received"], Decimal("77.00"))
        self.assertEqual(meituan["suggested_commission"], Decimal("0.00"))
        self.assertEqual(meituan["def_hexiao"], Decimal("64.68"))
        self.assertEqual(meituan["def_jinying"], Decimal("66.99"))

    def test_nanyang_douyin_received_is_order_received(self):
        parsed = ticket_ledger.parse_reconciliation(
            self._target_scenic_workbook(),
            "南阳森林野生动物世界8.1-8.1.xlsx",
            scenic_id="nanyang-wildlife",
            rate_hexiao=Decimal("0.80"),
            rate_settle=Decimal("0.85"),
            commission_rate=Decimal("0"),
            commission_override=Decimal("0"),
            ticket_product="南阳森林野生动物世界",
        )
        douyin = next(item for item in parsed["platforms"] if item["platform"] == "抖音")
        self.assertEqual(douyin["supplier_received"], Decimal("100.00"))
        self.assertEqual(douyin["suggested_commission"], Decimal("0.00"))
        self.assertEqual(douyin["def_hexiao"], Decimal("80.00"))
        self.assertEqual(douyin["def_jinying"], Decimal("85.00"))

    def test_zunyi_meituan_requires_technical_service_fee(self):
        with self.assertRaisesRegex(ValueError, "技术服务费"):
            ticket_ledger.parse_reconciliation(
                self._target_scenic_workbook(include_meituan_tech_fee=False),
                "遵义动物园8.1-8.1.xlsx",
                scenic_id="zunyi-zoo",
            )


if __name__ == "__main__":
    unittest.main()
