import json
import unittest
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import Workbook

from app.api.v1.endpoints import hotel_ledger as hotel_endpoint
from app.services.hotel_ledger import parse_hotel_file


RATES = {
    "scenic_id": "quancheng-ouleb",
    "rate_hexiao": Decimal("0.90"),
    "rate_settle": Decimal("0.95"),
    "commission_rate": Decimal("0.06"),
}


class HotelBrandPlatformParserTest(unittest.TestCase):
    @staticmethod
    def _workbook(sheet_titles: list[str]) -> bytes:
        workbook = Workbook()
        workbook.remove(workbook.active)
        for title in sheet_titles:
            worksheet = workbook.create_sheet(title)
            if "美团" in title:
                worksheet.append(["结算金额", "间夜", "入住日期", "离店日期"])
                worksheet.append([100, 1, datetime(2026, 1, 25), datetime(2026, 1, 26)])
            else:
                worksheet.append(["结算价", "间夜", "入住日期", "离店日期"])
                worksheet.append([100, 1, datetime(2026, 1, 25), datetime(2026, 1, 26)])
        output = BytesIO()
        workbook.save(output)
        workbook.close()
        return output.getvalue()

    def _branded_workbook(self) -> bytes:
        return self._workbook([
            "海洋携程1.25-2.21", "海洋美团1.25-2.21",
            "骑士携程1.25-2.21", "骑士美团1.25-2.21",
            "长颈鹿携程1.25-2.21", "长颈鹿美团1.25-2.21",
        ])

    def _legacy_workbook(self) -> bytes:
        return self._workbook(["携程1.25-1.31", "携程补充1.25-1.31"])

    def test_three_hotel_brands_are_split_into_six_rows(self):
        parsed = parse_hotel_file(
            self._branded_workbook(), "酒店2026.1.25-2.21.xlsx", **RATES
        )
        self.assertEqual(
            [(row["hotel_name"], row["platform"]) for row in parsed["platforms"]],
            [
                ("海洋", "携程"), ("海洋", "美团"), ("骑士", "携程"),
                ("骑士", "美团"), ("长颈鹿", "携程"), ("长颈鹿", "美团"),
            ],
        )

    def test_unbranded_sheets_still_merge_by_platform(self):
        parsed = parse_hotel_file(
            self._legacy_workbook(), "酒店2026.1.25-1.31.xlsx", **RATES
        )
        self.assertEqual(
            [(row["hotel_name"], row["platform"]) for row in parsed["platforms"]],
            [("", "携程")],
        )

    def test_branded_platform_merges_sheets_with_negative_cancellation(self):
        workbook = Workbook()
        workbook.remove(workbook.active)
        for title, amount, nights, stay_date in [
            ("海洋携程1.25-1.31", 100, 2, datetime(2026, 1, 25)),
            ("海洋携程补充1.25-1.31", -20, 1, datetime(2026, 1, 26)),
        ]:
            worksheet = workbook.create_sheet(title)
            worksheet.append(["结算价", "间夜", "入住日期", "离店日期"])
            worksheet.append([amount, nights, stay_date, stay_date])
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        parsed = parse_hotel_file(
            output.getvalue(), "酒店2026.1.25-1.31.xlsx", **RATES
        )
        platform = parsed["platforms"][0]

        self.assertEqual((platform["hotel_name"], platform["platform"]), ("海洋", "携程"))
        self.assertEqual(platform["base_received"], Decimal("80"))
        self.assertEqual(platform["room_nights"], 3)
        self.assertEqual(platform["order_count"], 2)
        self.assertEqual(platform["positive_count"], 1)
        self.assertEqual(
            [(day["b"], day["n"]) for day in json.loads(platform["daily_json"])],
            [("100", 2), ("-20", 1)],
        )

    def test_daily_recovery_matches_hotel_name_and_platform(self):
        row = SimpleNamespace(
            platform="携程", hotel_name="骑士", daily_json="",
            scenic_id="quancheng-ouleb", detail_stored="source.xlsx",
            detail_name="酒店2026.1.25-2.21.xlsx", source_file="source.xlsx",
            rate_hexiao=Decimal("0.90"), rate_settle=Decimal("0.94"),
            commission_rate=Decimal("0.06"),
        )
        parsed = {"platforms": [
            {"platform": "携程", "hotel_name": "海洋", "daily_json": "ocean"},
            {"platform": "携程", "hotel_name": "骑士", "daily_json": "knight"},
        ]}
        with TemporaryDirectory() as temp_dir:
            Path(temp_dir, "source.xlsx").write_bytes(b"xlsx")
            with patch.object(hotel_endpoint, "_detail_dir", return_value=Path(temp_dir)), \
                 patch.object(hotel_endpoint.hl_svc, "parse_hotel_file", return_value=parsed):
                self.assertEqual(hotel_endpoint._recover_daily_json(row), "knight")

    def test_daily_recovery_uses_unique_unbranded_platform_fallback(self):
        row = SimpleNamespace(
            platform="携程", hotel_name="默认酒店", daily_json="",
            scenic_id="quancheng-ouleb", detail_stored="source.xlsx",
            detail_name="酒店2026.1.25-1.31.xlsx", source_file="source.xlsx",
            rate_hexiao=Decimal("0.90"), rate_settle=Decimal("0.94"),
            commission_rate=Decimal("0.06"),
        )
        parsed = {"platforms": [
            {"platform": "携程", "hotel_name": "", "daily_json": "legacy"},
        ]}
        with TemporaryDirectory() as temp_dir:
            Path(temp_dir, "source.xlsx").write_bytes(b"xlsx")
            with patch.object(hotel_endpoint, "_detail_dir", return_value=Path(temp_dir)), \
                 patch.object(hotel_endpoint.hl_svc, "parse_hotel_file", return_value=parsed):
                self.assertEqual(hotel_endpoint._recover_daily_json(row), "legacy")
