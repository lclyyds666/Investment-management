import unittest
from datetime import date, timedelta
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import patch

import openpyxl
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

import app.db.init_db  # noqa: F401
from app.api.deps import get_current_user
from app.core.enums import (
    AssignmentStatus,
    ContractStatus,
    ContractType,
    WorkflowAction,
    WorkflowTaskStatus,
)
from app.db.base import Base
from app.db.session import get_db
from app.main import create_app
from app.models.approval import Approval
from app.models.contract import Contract
from app.models.approval_form import ApprovalForm, ApprovalFormAction
from app.models.organization import ExternalAssignment, Organization, Position, UserAssignment
from app.models.user import User
from app.models.workflow import WorkflowInstance, WorkflowTask, WorkflowTaskAction
from app.services.approval_print import build_approval_form_xlsx
from app.services.organization_catalog import seed_authorization_catalog
from app.services.workflow_engine import seed_workflow_definitions, start_workflow


class WorkflowApiTest(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(self.engine)
        self.db = Session(self.engine)
        seed_authorization_catalog(self.db)
        self.publisher = self.add_user("publisher", "Publisher")
        self.handler = self.add_user("handler", "Handler")
        self.reviewer_a = self.add_user("reviewer-a", "Reviewer A")
        self.reviewer_b = self.add_user("reviewer-b", "Reviewer B")
        self.leader = self.add_user("leader", "Leader")
        self.other_leader = self.add_user("other-leader", "Other Leader")
        self.legal = self.add_user("legal", "Legal")
        self.risk = self.add_user("risk", "Risk")
        self.governance = self.add_user("governance", "Governance")
        self.admin = self.add_user("admin", "Information Maintainer", is_superuser=True)
        self.assign(self.handler, "supplymanagement", "supply.business_handler")
        self.assign(self.reviewer_a, "supplymanagement", "supply.business_reviewer")
        self.assign(self.reviewer_b, "supplymanagement", "supply.business_reviewer")
        self.assign(self.leader, "supplymanagement", "supply.company_leader")
        self.assign(self.other_leader, "supplymanagement", "supply.company_leader")
        legal_assignment = self.assign(self.legal, "external.legal", "external.legal_counsel")
        self.db.add(ExternalAssignment(
            assignment_id=legal_assignment.id,
            provider_name="Legal Firm",
            service_scopes=["contract_legal_review"],
        ))
        self.assign(self.risk, "investment.legal_risk", "investment.duty.supply_risk_review")
        self.assign(self.governance, "supplymanagement", "governance.supply_leader")
        seed_workflow_definitions(self.db, self.publisher.id)
        self.db.commit()
        self.contract = Contract(
            contract_no="WF-API-1",
            title="Workflow API Contract",
            status=ContractStatus.DRAFT,
            created_by=self.handler.id,
        )
        self.db.add(self.contract)
        self.db.commit()
        self.instance = start_workflow(
            self.db,
            "contract",
            self.contract.id,
            self.handler,
            {
                "company_leader": self.leader.id,
                "legal_counsel": self.legal.id,
                "supply_governance_leader": self.governance.id,
            },
        )
        self.db.commit()
        self.app = create_app()
        self.app.dependency_overrides[get_db] = lambda: self.db
        self.current_user = self.handler
        self.app.dependency_overrides[get_current_user] = lambda: self.current_user
        self.client = TestClient(self.app)

    def tearDown(self):
        self.client.close()
        self.app.dependency_overrides.clear()
        self.db.close()
        self.engine.dispose()

    def add_user(self, username, full_name, *, is_superuser=False):
        user = User(
            username=username,
            full_name=full_name,
            hashed_password="test",
            is_superuser=is_superuser,
            is_active=True,
        )
        self.db.add(user)
        self.db.flush()
        return user

    def assign(self, user, organization_code, position_code, *, valid_until=None, status=AssignmentStatus.ACTIVE):
        assignment = UserAssignment(
            user_id=user.id,
            organization_id=self.db.scalar(select(Organization.id).where(Organization.code == organization_code)),
            position_id=self.db.scalar(select(Position.id).where(Position.code == position_code)),
            valid_from=date(2026, 1, 1),
            valid_until=valid_until,
            status=status,
        )
        self.db.add(assignment)
        self.db.flush()
        return assignment

    def add_external_legal(self, username, *, service_scopes=None):
        user = self.add_user(username, username)
        assignment = self.assign(user, "external.legal", "external.legal_counsel")
        self.db.add(ExternalAssignment(
            assignment_id=assignment.id,
            provider_name="Legal Firm",
            service_scopes=(
                ["contract_legal_review"]
                if service_scopes is None
                else service_scopes
            ),
        ))
        self.db.flush()
        return user, assignment

    def advance_contract_to_legal(self):
        self.current_user = self.leader
        response = self.client.post(
            f"/api/v1/workflows/tasks/{self.active_task().id}/approve",
            json={"comment": "负责人同意"},
        )
        self.assertEqual(response.status_code, 200, response.text)

    def active_task(self):
        return self.db.scalar(
            select(WorkflowTask).where(
                WorkflowTask.instance_id == self.instance.id,
                WorkflowTask.status == WorkflowTaskStatus.ACTIVE,
            )
        )

    def test_candidates_require_submit_permission_and_filter_ineffective_assignments(self):
        expired = self.add_user("expired-leader", "Expired Leader")
        self.assign(expired, "supplymanagement", "supply.company_leader", valid_until=date.today() - timedelta(days=1))
        inactive = self.add_user("inactive-leader", "Inactive Leader")
        self.assign(inactive, "supplymanagement", "supply.company_leader", status=AssignmentStatus.INACTIVE)
        self.db.commit()

        response = self.client.get(
            "/api/v1/workflows/candidates",
            params={"workflow_code": "supply.contract.v2", "node_code": "company_leader"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["code"], 0)
        self.assertEqual(
            {item["user_id"] for item in response.json()["data"]},
            {self.leader.id, self.other_leader.id},
        )
        self.current_user = self.reviewer_a
        denied = self.client.get(
            "/api/v1/workflows/candidates",
            params={"workflow_code": "supply.contract.v2", "node_code": "company_leader"},
        )
        self.assertEqual(denied.status_code, 403)

    def test_inbox_exposes_shared_task_to_all_position_holders(self):
        payment = ApprovalForm(
            form_type="payment",
            contract_no="WF-API-PAYMENT",
            business_desc="Shared Inbox",
            status=ContractStatus.DRAFT,
            created_by=self.handler.id,
        )
        self.db.add(payment)
        self.db.commit()
        instance = start_workflow(
            self.db,
            "payment_approval",
            payment.id,
            self.handler,
            {
                "company_leader": self.leader.id,
                "supply_governance_leader": self.governance.id,
            },
        )
        self.db.commit()
        self.current_user = self.reviewer_a
        first = self.client.get("/api/v1/workflows/my-tasks")
        self.current_user = self.reviewer_b
        second = self.client.get("/api/v1/workflows/my-tasks")

        self.assertEqual(first.status_code, 200)
        self.assertEqual(second.status_code, 200)
        self.assertEqual(first.json()["data"][0]["instance_id"], instance.id)
        self.assertEqual(second.json()["data"][0]["instance_id"], instance.id)
        self.assertIn("approve", first.json()["data"][0]["allowed_actions"])

    def test_pending_count_allows_assigned_legal_and_does_not_leak_unassigned_target(self):
        other_legal, _ = self.add_external_legal("other-legal")
        self.db.commit()
        self.advance_contract_to_legal()

        self.current_user = self.legal
        assigned = self.client.get("/api/v1/approval/pending-count")
        self.current_user = other_legal
        unassigned = self.client.get("/api/v1/approval/pending-count")
        self.current_user = self.publisher
        unauthorized = self.client.get("/api/v1/approval/pending-count")

        self.assertEqual(assigned.status_code, 200, assigned.text)
        self.assertEqual(assigned.json()["data"], {
            "contract": 1,
            "business": 0,
            "total": 1,
        })
        self.assertEqual(unassigned.status_code, 200, unassigned.text)
        self.assertEqual(unassigned.json()["data"], {
            "contract": 0,
            "business": 0,
            "total": 0,
        })
        self.assertEqual(unauthorized.status_code, 403)

    def test_pending_count_excludes_designated_legal_without_required_scope(self):
        self.advance_contract_to_legal()
        legal_detail = self.db.scalar(select(ExternalAssignment).where(
            ExternalAssignment.assignment_id == self.active_task().designated_assignment_id
        ))
        legal_detail.service_scopes = []
        self.db.commit()
        self.current_user = self.legal

        response = self.client.get("/api/v1/approval/pending-count")

        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.json()["data"]["contract"], 0)
        self.db.expire_all()
        self.assertEqual(self.active_task(), None)

    def test_pending_count_excludes_expired_locked_legal_assignment(self):
        self.advance_contract_to_legal()
        locked_assignment = self.db.get(
            UserAssignment,
            self.active_task().designated_assignment_id,
        )
        locked_assignment.valid_until = date.today() - timedelta(days=1)
        replacement = self.assign(self.legal, "external.legal", "external.legal_counsel")
        self.db.add(ExternalAssignment(
            assignment_id=replacement.id,
            provider_name="Replacement Legal Firm",
            service_scopes=["contract_legal_review"],
        ))
        self.db.commit()
        self.current_user = self.legal

        response = self.client.get("/api/v1/approval/pending-count")

        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.json()["data"]["contract"], 0)
        self.assertEqual(
            self.db.scalar(select(WorkflowTask.status).where(
                WorkflowTask.instance_id == self.instance.id,
                WorkflowTask.sequence == 2,
            )),
            WorkflowTaskStatus.AWAITING_REASSIGNMENT,
        )

    def test_designated_task_is_visible_and_actionable_only_by_selected_user(self):
        task = self.active_task()
        self.current_user = self.other_leader
        inbox = self.client.get("/api/v1/workflows/my-tasks")
        denied = self.client.post(f"/api/v1/workflows/tasks/{task.id}/approve", json={"comment": "no"})

        self.assertEqual(inbox.json()["data"], [])
        self.assertEqual(denied.status_code, 403)
        self.current_user = self.leader
        approved = self.client.post(f"/api/v1/workflows/tasks/{task.id}/approve", json={"comment": "ok"})
        self.assertEqual(approved.status_code, 200)

    def test_superuser_cannot_approve_and_second_action_returns_actor_snapshot(self):
        task = self.active_task()
        self.current_user = self.admin
        self.assertEqual(
            self.client.post(f"/api/v1/workflows/tasks/{task.id}/approve", json={"comment": "admin"}).status_code,
            403,
        )
        self.current_user = self.leader
        self.assertEqual(
            self.client.post(f"/api/v1/workflows/tasks/{task.id}/approve", json={"comment": "done"}).status_code,
            200,
        )
        self.current_user = self.other_leader
        conflict = self.client.post(f"/api/v1/workflows/tasks/{task.id}/approve", json={"comment": "late"})

        self.assertEqual(conflict.status_code, 409)
        self.assertEqual(conflict.json()["detail"]["actor"], self.leader.full_name)
        self.assertEqual(conflict.json()["detail"]["action"], WorkflowAction.APPROVE.value)
        self.assertIsNotNone(conflict.json()["detail"]["completed_at"])
        self.assertEqual(
            self.db.scalar(select(WorkflowTaskAction.actor_name).where(WorkflowTaskAction.task_id == task.id)),
            self.leader.full_name,
        )
        self.current_user = self.handler
        self.assertEqual(
            self.client.get(
                "/api/v1/workflows/candidates",
                params={"workflow_code": "supply.contract.v2", "node_code": "company_leader"},
            ).status_code,
            200,
        )

    def test_reject_requires_nonblank_reason_and_timeline_preserves_snapshot(self):
        task = self.active_task()
        self.current_user = self.leader
        blank = self.client.post(f"/api/v1/workflows/tasks/{task.id}/reject", json={"reason": "   "})
        self.assertEqual(blank.status_code, 422)
        rejected = self.client.post(f"/api/v1/workflows/tasks/{task.id}/reject", json={"reason": "补充材料"})
        self.assertEqual(rejected.status_code, 200)
        timeline = self.client.get(f"/api/v1/workflows/instances/{self.instance.id}/timeline")
        self.assertEqual(timeline.status_code, 200)
        action = next(item for item in timeline.json()["data"] if item["action"] == "return")
        self.assertEqual(action["comment"], "补充材料")
        self.assertEqual(action["actor_name"], self.leader.full_name)

    def test_generic_workflow_actions_project_contract_approval_snapshots(self):
        submit_projection = self.db.scalar(
            select(Approval).where(Approval.contract_id == self.contract.id)
        )
        self.assertIsNotNone(submit_projection.workflow_task_action_id)
        self.assertEqual(submit_projection.approver_role, "supply.business_handler")
        self.assertEqual(submit_projection.position_code, "supply.business_handler")
        self.assertEqual(submit_projection.position_name, "业务经办")

        task = self.active_task()
        self.current_user = self.leader
        approved = self.client.post(
            f"/api/v1/workflows/tasks/{task.id}/approve",
            json={"comment": "同意"},
        )

        self.assertEqual(approved.status_code, 200)
        projection = self.db.scalar(
            select(Approval)
            .where(Approval.contract_id == self.contract.id)
            .order_by(Approval.id.desc())
        )
        self.assertEqual(projection.action.value, "approve")
        self.assertEqual(projection.position_code, "supply.company_leader")
        self.assertEqual(projection.position_name, "供应链公司负责人")
        self.assertEqual(projection.organization_code, "supplymanagement")

        legal_task = self.active_task()
        self.current_user = self.legal
        returned = self.client.post(
            f"/api/v1/workflows/tasks/{legal_task.id}/reject",
            json={"reason": "请补充"},
        )

        self.assertEqual(returned.status_code, 200)
        projection = self.db.scalar(
            select(Approval)
            .where(Approval.contract_id == self.contract.id)
            .order_by(Approval.id.desc())
        )
        self.assertEqual(projection.action.value, "reject")
        self.assertEqual(projection.position_code, "external.legal_counsel")


    def test_timeline_requires_target_view_permission_but_allows_information_maintainer(self):
        self.current_user = self.handler
        business_reader = self.client.get(f"/api/v1/workflows/instances/{self.instance.id}/timeline")
        self.assertEqual(business_reader.status_code, 200)
        self.current_user = self.publisher
        denied = self.client.get(f"/api/v1/workflows/instances/{self.instance.id}/timeline")
        self.assertEqual(denied.status_code, 403)
        self.current_user = self.admin
        allowed = self.client.get(f"/api/v1/workflows/instances/{self.instance.id}/timeline")
        self.assertEqual(allowed.status_code, 200)

    def test_reassignment_requires_superuser_reason_and_exact_effective_position(self):
        task = self.active_task()
        self.current_user = self.handler
        denied = self.client.post(
            f"/api/v1/workflows/tasks/{task.id}/reassign",
            json={"user_id": self.other_leader.id, "reason": "人员调整"},
        )
        self.assertEqual(denied.status_code, 403)
        self.current_user = self.admin
        blank = self.client.post(
            f"/api/v1/workflows/tasks/{task.id}/reassign",
            json={"user_id": self.other_leader.id, "reason": "   "},
        )
        self.assertEqual(blank.status_code, 422)
        wrong_position = self.client.post(
            f"/api/v1/workflows/tasks/{task.id}/reassign",
            json={"user_id": self.reviewer_a.id, "reason": "人员调整"},
        )
        self.assertEqual(wrong_position.status_code, 422)

        with patch.object(self.db, "commit", wraps=self.db.commit) as commit:
            response = self.client.post(
                f"/api/v1/workflows/tasks/{task.id}/reassign",
                json={"user_id": self.other_leader.id, "reason": "人员调整"},
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(commit.call_count, 1)
        self.db.expire_all()
        persisted = self.db.get(WorkflowTask, task.id)
        self.assertEqual(persisted.designated_user_id, self.other_leader.id)
        self.assertEqual(persisted.status, WorkflowTaskStatus.ACTIVE)
        snapshot = self.db.scalar(
            select(WorkflowTaskAction)
            .where(WorkflowTaskAction.task_id == task.id, WorkflowTaskAction.action == WorkflowAction.REASSIGN)
        )
        self.assertEqual(snapshot.previous_assignee_id, self.leader.id)
        self.assertEqual(snapshot.new_assignee_id, self.other_leader.id)
        self.assertEqual(snapshot.previous_assignee_name, "Leader")
        self.assertEqual(snapshot.new_assignee_name, "Other Leader")
        self.assertEqual(snapshot.reason, "人员调整")
        self.assertEqual(snapshot.actor_name, self.admin.full_name)
        self.leader.full_name = "Renamed Former Leader"
        self.other_leader.full_name = "Renamed New Leader"
        snapshot.previous_assignee_id = None
        snapshot.new_assignee_id = None
        self.db.commit()
        self.current_user = self.handler
        timeline = self.client.get(f"/api/v1/workflows/instances/{self.instance.id}/timeline")
        reassign_snapshot = next(
            item for item in timeline.json()["data"] if item["action"] == WorkflowAction.REASSIGN.value
        )
        self.assertIsNone(reassign_snapshot["previous_assignee_id"])
        self.assertIsNone(reassign_snapshot["new_assignee_id"])
        self.assertEqual(reassign_snapshot["previous_assignee_name"], "Leader")
        self.assertEqual(reassign_snapshot["new_assignee_name"], "Other Leader")

    def test_reassignment_cas_conflict_rolls_back_task_and_action_atomically(self):
        task = self.active_task()
        original = {
            "designated_user_id": task.designated_user_id,
            "designated_assignment_id": task.designated_assignment_id,
            "status": task.status,
            "version": task.version,
        }
        self.current_user = self.admin
        real_execute = self.db.execute

        def stale_task_update(statement, *args, **kwargs):
            result = real_execute(statement, *args, **kwargs)
            if getattr(statement, "is_update", False) and statement.table.name == "wf_task":
                return SimpleNamespace(rowcount=0)
            return result

        with patch.object(self.db, "execute", side_effect=stale_task_update):
            response = self.client.post(
                f"/api/v1/workflows/tasks/{task.id}/reassign",
                json={"user_id": self.other_leader.id, "reason": "并发改派"},
            )

        self.assertEqual(response.status_code, 409)
        self.assertEqual(response.json()["detail"]["code"], "workflow_task_reassignment_conflict")
        self.db.expire_all()
        persisted = self.db.get(WorkflowTask, task.id)
        self.assertEqual(persisted.designated_user_id, original["designated_user_id"])
        self.assertEqual(persisted.designated_assignment_id, original["designated_assignment_id"])
        self.assertEqual(persisted.status, original["status"])
        self.assertEqual(persisted.version, original["version"])
        self.assertIsNone(self.db.scalar(
            select(WorkflowTaskAction).where(
                WorkflowTaskAction.task_id == task.id,
                WorkflowTaskAction.action == WorkflowAction.REASSIGN,
            )
        ))

    def test_reassignment_rejects_duplicate_person_and_shared_task(self):
        task = self.active_task()
        self.current_user = self.admin
        duplicate = self.client.post(
            f"/api/v1/workflows/tasks/{task.id}/reassign",
            json={"user_id": self.handler.id, "reason": "重复人员"},
        )
        self.assertEqual(duplicate.status_code, 422)

        task.assignee_mode = "shared_position"
        self.db.commit()
        shared = self.client.post(
            f"/api/v1/workflows/tasks/{task.id}/reassign",
            json={"user_id": self.other_leader.id, "reason": "错误模式"},
        )
        self.assertEqual(shared.status_code, 422)

    def test_reassignment_reactivates_invalid_designated_task_without_automatic_replacement(self):
        task = self.active_task()
        old_assignment = self.db.get(UserAssignment, task.designated_assignment_id)
        old_assignment.valid_until = date.today() - timedelta(days=1)
        self.db.commit()
        self.current_user = self.leader

        inbox = self.client.get("/api/v1/workflows/my-tasks")

        self.assertEqual(inbox.json()["data"], [])
        self.db.expire_all()
        self.assertEqual(self.db.get(WorkflowTask, task.id).status, WorkflowTaskStatus.AWAITING_REASSIGNMENT)
        self.assertEqual(self.db.get(WorkflowTask, task.id).designated_user_id, self.leader.id)
        self.current_user = self.admin
        response = self.client.post(
            f"/api/v1/workflows/tasks/{task.id}/reassign",
            json={"user_id": self.other_leader.id, "reason": "原任职失效"},
        )
        self.assertEqual(response.status_code, 200)
        self.db.expire_all()
        persisted = self.db.get(WorkflowTask, task.id)
        self.assertEqual(persisted.status, WorkflowTaskStatus.ACTIVE)
        self.assertEqual(persisted.designated_user_id, self.other_leader.id)


class ContractWorkflowApiTest(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(self.engine)
        self.db = Session(self.engine)
        seed_authorization_catalog(self.db)
        self.publisher = self.add_user("publisher", "Publisher")
        self.handler = self.add_user("contract-handler", "Contract Handler")
        self.leader = self.add_user("contract-leader", "Contract Leader")
        self.legal = self.add_user("contract-legal", "Contract Legal")
        self.risk = self.add_user("contract-risk", "Contract Risk")
        self.governance = self.add_user("contract-governance", "Contract Governance")
        self.other_handler = self.add_user("contract-other-handler", "Other Handler")
        self.assign(self.handler, "supplymanagement", "supply.business_handler")
        self.assign(self.other_handler, "supplymanagement", "supply.business_handler")
        self.assign(self.leader, "supplymanagement", "supply.company_leader")
        legal_assignment = self.assign(self.legal, "external.legal", "external.legal_counsel")
        self.db.add(ExternalAssignment(
            assignment_id=legal_assignment.id,
            provider_name="Legal Firm",
            service_scopes=["contract_legal_review"],
        ))
        self.assign(self.risk, "investment.legal_risk", "investment.duty.supply_risk_review")
        self.assign(self.governance, "supplymanagement", "governance.supply_leader")
        seed_workflow_definitions(self.db, self.publisher.id)
        self.contract = Contract(
            contract_no="WF-CONTRACT-1",
            title="Contract Route Cutover",
            status=ContractStatus.DRAFT,
            created_by=self.handler.id,
        )
        self.db.add(self.contract)
        self.db.commit()
        self.current_user = self.handler
        self.app = create_app()
        self.app.dependency_overrides[get_db] = lambda: self.db
        self.app.dependency_overrides[get_current_user] = lambda: self.current_user
        self.client = TestClient(self.app)

    def tearDown(self):
        self.client.close()
        self.app.dependency_overrides.clear()
        self.db.close()
        self.engine.dispose()

    def add_user(self, username, full_name):
        user = User(
            username=username,
            full_name=full_name,
            hashed_password="test",
            is_active=True,
        )
        self.db.add(user)
        self.db.flush()
        return user

    def assign(self, user, organization_code, position_code):
        assignment = UserAssignment(
            user_id=user.id,
            organization_id=self.db.scalar(
                select(Organization.id).where(Organization.code == organization_code)
            ),
            position_id=self.db.scalar(
                select(Position.id).where(Position.code == position_code)
            ),
            valid_from=date(2026, 1, 1),
            status=AssignmentStatus.ACTIVE,
        )
        self.db.add(assignment)
        self.db.flush()
        return assignment

    def designation_payload(self):
        return {
            "designated_users": {
                "company_leader": self.leader.id,
                "legal_counsel": self.legal.id,
                "supply_governance_leader": self.governance.id,
            }
        }

    def test_contract_submit_requires_three_designated_users(self):
        response = self.client.post(
            f"/api/v1/contracts/{self.contract.id}/submit",
            json={"designated_users": {"company_leader": self.leader.id}},
        )
        self.assertEqual(response.status_code, 422)

    def test_contract_chain_uses_supply_governance_leader(self):
        response = self.client.post(
            f"/api/v1/contracts/{self.contract.id}/submit",
            json=self.designation_payload(),
        )
        self.assertEqual(response.status_code, 200, response.text)
        data = response.json()["data"]
        self.assertIsNotNone(data["workflow_instance_id"])
        self.assertEqual(data["active_task"]["position_code"], "supply.company_leader")
        self.assertFalse(data["can_act"])
        tasks = list(self.db.scalars(
            select(WorkflowTask)
            .where(WorkflowTask.instance_id == data["workflow_instance_id"])
            .order_by(WorkflowTask.sequence)
        ))
        self.assertEqual(tasks[-1].required_position_code, "governance.supply_leader")

    def test_todo_and_legacy_wrappers_use_current_workflow_task(self):
        submitted = self.client.post(
            f"/api/v1/contracts/{self.contract.id}/submit",
            json=self.designation_payload(),
        )
        self.assertEqual(submitted.status_code, 200, submitted.text)
        self.current_user = self.leader
        todo = self.client.get("/api/v1/contracts/todo")
        self.assertEqual([item["id"] for item in todo.json()["data"]], [self.contract.id])
        self.assertTrue(todo.json()["data"][0]["can_act"])
        approved = self.client.post(
            f"/api/v1/contracts/{self.contract.id}/approve",
            json={"comment": "负责人同意"},
        )
        self.assertEqual(approved.status_code, 200, approved.text)
        self.assertEqual(approved.json()["data"]["active_task"]["position_code"], "external.legal_counsel")
        self.current_user = self.legal
        returned = self.client.post(
            f"/api/v1/contracts/{self.contract.id}/reject",
            json={"comment": "退回补充"},
        )
        self.assertEqual(returned.status_code, 200, returned.text)
        self.assertEqual(returned.json()["data"]["active_task"]["position_code"], "supply.company_leader")
        self.current_user = self.handler
        timeline = self.client.get(
            f"/api/v1/workflows/instances/{self.contract.workflow_instance_id}/timeline"
        ).json()["data"]
        self.assertEqual(timeline[-1]["action"], "return")

    def test_return_to_handler_resubmits_same_instance_without_designations(self):
        first = self.client.post(
            f"/api/v1/contracts/{self.contract.id}/submit",
            json=self.designation_payload(),
        )
        self.assertEqual(first.status_code, 200, first.text)
        instance_id = first.json()["data"]["workflow_instance_id"]
        self.current_user = self.leader
        returned = self.client.post(
            f"/api/v1/contracts/{self.contract.id}/reject",
            json={"comment": "请修改合同"},
        )
        self.assertEqual(returned.status_code, 200, returned.text)
        self.current_user = self.handler
        updated = self.client.put(
            f"/api/v1/contracts/{self.contract.id}",
            json={"remark": "已按意见修改"},
        )
        self.assertEqual(updated.status_code, 200, updated.text)

        resumed = self.client.post(
            f"/api/v1/contracts/{self.contract.id}/submit",
            json={},
        )

        self.assertEqual(resumed.status_code, 200, resumed.text)
        self.assertEqual(resumed.json()["data"]["workflow_instance_id"], instance_id)
        self.assertEqual(
            resumed.json()["data"]["active_task"]["position_code"],
            "supply.company_leader",
        )
        actions = list(self.db.scalars(
            select(WorkflowTaskAction)
            .join(WorkflowTaskAction.task)
            .where(WorkflowTask.instance_id == instance_id)
            .order_by(WorkflowTaskAction.id)
        ))
        self.assertEqual(
            [action.action for action in actions],
            [WorkflowAction.SUBMIT, WorkflowAction.RETURN, WorkflowAction.SUBMIT],
        )
        approvals = list(self.db.scalars(
            select(Approval)
            .where(Approval.contract_id == self.contract.id)
            .order_by(Approval.id)
        ))
        self.assertEqual([item.action.value for item in approvals], ["approve", "reject", "approve"])
        self.assertEqual(self.db.query(WorkflowTask).filter(
            WorkflowTask.instance_id == instance_id,
            WorkflowTask.status == WorkflowTaskStatus.ACTIVE,
        ).one().sequence, 1)

    def test_resubmit_rejects_non_submitter_and_non_handler_active_task(self):
        first = self.client.post(
            f"/api/v1/contracts/{self.contract.id}/submit",
            json=self.designation_payload(),
        )
        self.assertEqual(first.status_code, 200, first.text)
        self.current_user = self.other_handler
        non_submitter = self.client.post(
            f"/api/v1/contracts/{self.contract.id}/submit",
            json={},
        )
        self.assertEqual(non_submitter.status_code, 403)
        self.current_user = self.handler
        wrong_active_task = self.client.post(
            f"/api/v1/contracts/{self.contract.id}/submit",
            json={},
        )
        self.assertEqual(wrong_active_task.status_code, 422)

    def test_contract_reassignment_stays_timeline_only(self):
        second_leader = self.add_user("second-contract-leader", "Second Contract Leader")
        self.assign(second_leader, "supplymanagement", "supply.company_leader")
        self.db.commit()
        first = self.client.post(
            f"/api/v1/contracts/{self.contract.id}/submit",
            json=self.designation_payload(),
        )
        task_id = first.json()["data"]["active_task"]["id"]
        approvals_before = self.db.query(Approval).count()
        admin = self.add_user("contract-admin", "Contract Admin")
        admin.is_superuser = True
        self.db.commit()
        self.current_user = admin

        response = self.client.post(
            f"/api/v1/workflows/tasks/{task_id}/reassign",
            json={"user_id": second_leader.id, "reason": "负责人调整"},
        )

        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(self.db.query(Approval).count(), approvals_before)
        reassignment = self.db.scalar(select(WorkflowTaskAction).where(
            WorkflowTaskAction.task_id == task_id,
            WorkflowTaskAction.action == WorkflowAction.REASSIGN,
        ))
        self.assertIsNotNone(reassignment)
        self.assertIsNone(self.db.scalar(select(Approval).where(
            Approval.workflow_task_action_id == reassignment.id,
        )))

    def test_legacy_reject_wrapper_requires_nonblank_reason(self):
        submitted = self.client.post(
            f"/api/v1/contracts/{self.contract.id}/submit",
            json=self.designation_payload(),
        )
        self.assertEqual(submitted.status_code, 200, submitted.text)
        self.current_user = self.leader

        response = self.client.post(
            f"/api/v1/contracts/{self.contract.id}/reject",
            json={"comment": "   "},
        )

        self.assertEqual(response.status_code, 422)

    def test_external_legal_counsel_sees_only_designated_contract(self):
        other = Contract(
            contract_no="WF-CONTRACT-OTHER",
            title="Other Contract",
            status=ContractStatus.DRAFT,
            created_by=self.handler.id,
        )
        self.db.add(other)
        self.db.commit()
        submitted = self.client.post(
            f"/api/v1/contracts/{self.contract.id}/submit",
            json=self.designation_payload(),
        )
        self.assertEqual(submitted.status_code, 200, submitted.text)

        self.current_user = self.legal
        listing = self.client.get("/api/v1/contracts")
        assigned = self.client.get(f"/api/v1/contracts/{self.contract.id}")
        unassigned = self.client.get(f"/api/v1/contracts/{other.id}")

        self.assertEqual(listing.status_code, 200, listing.text)
        self.assertEqual([item["id"] for item in listing.json()["data"]], [self.contract.id])
        self.assertEqual(assigned.status_code, 200)
        self.assertEqual(unassigned.status_code, 403)

    def test_external_legal_counsel_without_designation_sees_empty_list(self):
        self.current_user = self.legal

        response = self.client.get("/api/v1/contracts")

        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.json()["data"], [])

    def test_legal_doc_uses_position_snapshots_and_historical_fallback(self):
        self.db.add_all([
            Approval(
                contract_id=self.contract.id,
                approver_id=self.legal.id,
                approver_role="ignored_for_new_row",
                action="approve",
                comment="法律意见",
                position_code="external.legal_counsel",
                position_name="外聘法律顾问",
            ),
            Approval(
                contract_id=self.contract.id,
                approver_id=self.risk.id,
                approver_role="risk_auditor",
                action="approve",
                comment="风控意见",
                position_code="investment.duty.supply_risk_review",
                position_name="供应链风控复核",
            ),
            Approval(
                contract_id=self.contract.id,
                approver_id=self.governance.id,
                approver_role="invest_director",
                action="approve",
                comment="历史治理意见",
            ),
        ])
        self.db.commit()

        with patch(
            "app.api.v1.endpoints.contract.legal_doc_svc.build_legal_doc",
            return_value=b"docx",
        ) as build:
            response = self.client.get(f"/api/v1/contracts/{self.contract.id}/legal-doc")

        self.assertEqual(response.status_code, 200, response.text)
        opinions = build.call_args.args[1]
        self.assertEqual(opinions["external.legal_counsel"]["comment"], "法律意见")
        self.assertEqual(
            opinions["investment.duty.supply_risk_review"]["comment"],
            "风控意见",
        )
        self.assertEqual(opinions["governance.supply_leader"]["comment"], "历史治理意见")


class ApprovalFormWorkflowApiTest(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(self.engine)
        self.db = Session(self.engine)
        seed_authorization_catalog(self.db)
        self.publisher = self.add_user("form-publisher", "Publisher")
        self.handler = self.add_user("form-handler", "Form Handler")
        self.other_handler = self.add_user("form-other-handler", "Other Handler")
        self.reviewer_a = self.add_user("form-reviewer-a", "Reviewer A")
        self.reviewer_b = self.add_user("form-reviewer-b", "Reviewer B")
        self.finance_handler = self.add_user("form-finance-handler", "Finance Handler")
        self.leader = self.add_user("form-leader", "Company Leader")
        self.other_leader = self.add_user("form-other-leader", "Other Leader")
        self.risk = self.add_user("form-risk", "Risk Reviewer")
        self.finance_reviewer = self.add_user("form-finance-reviewer", "Finance Reviewer")
        self.governance = self.add_user("form-governance", "Governance Leader")
        self.legal = self.add_user("form-legal", "Legal Counsel")
        self.admin = self.add_user("form-admin", "Information Maintainer", is_superuser=True)
        self.assign(self.handler, "supplymanagement", "supply.business_handler")
        self.assign(self.other_handler, "supplymanagement", "supply.business_handler")
        self.assign(self.reviewer_a, "supplymanagement", "supply.business_reviewer")
        self.assign(self.reviewer_b, "supplymanagement", "supply.business_reviewer")
        self.assign(self.finance_handler, "supplymanagement", "supply.finance_handler")
        self.assign(self.leader, "supplymanagement", "supply.company_leader")
        self.assign(self.other_leader, "supplymanagement", "supply.company_leader")
        self.assign(self.risk, "investment.legal_risk", "investment.duty.supply_risk_review")
        self.assign(
            self.finance_reviewer,
            "investment.asset_finance",
            "investment.duty.supply_finance_review",
        )
        self.assign(self.governance, "supplymanagement", "governance.supply_leader")
        legal_assignment = self.assign(self.legal, "external.legal", "external.legal_counsel")
        self.db.add(ExternalAssignment(
            assignment_id=legal_assignment.id,
            provider_name="Legal Firm",
            service_scopes=["contract_legal_review"],
        ))
        seed_workflow_definitions(self.db, self.publisher.id)
        self.payment = self.add_form(ContractType.PAYMENT, "WF-FORM-PAYMENT")
        self.business = self.add_form(ContractType.BUSINESS, "WF-FORM-BUSINESS")
        self.db.commit()
        self.current_user = self.handler
        self.app = create_app()
        self.app.dependency_overrides[get_db] = lambda: self.db
        self.app.dependency_overrides[get_current_user] = lambda: self.current_user
        self.client = TestClient(self.app)

    def tearDown(self):
        self.client.close()
        self.app.dependency_overrides.clear()
        self.db.close()
        self.engine.dispose()

    def add_user(self, username, full_name, *, is_superuser=False):
        user = User(
            username=username,
            full_name=full_name,
            hashed_password="test",
            is_superuser=is_superuser,
            is_active=True,
        )
        self.db.add(user)
        self.db.flush()
        return user

    def assign(self, user, organization_code, position_code):
        assignment = UserAssignment(
            user_id=user.id,
            organization_id=self.db.scalar(
                select(Organization.id).where(Organization.code == organization_code)
            ),
            position_id=self.db.scalar(
                select(Position.id).where(Position.code == position_code)
            ),
            valid_from=date(2026, 1, 1),
            status=AssignmentStatus.ACTIVE,
        )
        self.db.add(assignment)
        self.db.flush()
        return assignment

    def add_form(self, form_type, contract_no):
        form = ApprovalForm(
            form_type=form_type,
            contract_no=contract_no,
            business_desc="Workflow cutover",
            status=ContractStatus.DRAFT,
            created_by=self.handler.id,
        )
        self.db.add(form)
        self.db.flush()
        return form

    def designation_payload(self):
        return {
            "designated_users": {
                "company_leader": self.leader.id,
                "supply_governance_leader": self.governance.id,
            }
        }

    def submit(self, form):
        return self.client.post(
            f"/api/v1/approval-forms/{form.id}/submit",
            json=self.designation_payload(),
        )

    def active_task(self, form):
        return self.db.scalar(select(WorkflowTask).where(
            WorkflowTask.instance_id == form.workflow_instance_id,
            WorkflowTask.status == WorkflowTaskStatus.ACTIVE,
        ))

    def test_payment_and_business_require_two_designations_and_materialize_confirmed_chains(self):
        missing_payment = self.add_form(ContractType.PAYMENT, "WF-FORM-MISSING-PAYMENT")
        missing_business = self.add_form(ContractType.BUSINESS, "WF-FORM-MISSING-BUSINESS")
        self.db.commit()
        for form in (missing_payment, missing_business):
            with self.subTest(form_type=form.form_type):
                missing = self.client.post(
                    f"/api/v1/approval-forms/{form.id}/submit",
                    json={"designated_users": {"company_leader": self.leader.id}},
                )
                self.assertEqual(missing.status_code, 422)

        payment = self.submit(self.payment)
        business = self.submit(self.business)

        self.assertEqual(payment.status_code, 200, payment.text)
        self.assertEqual(business.status_code, 200, business.text)
        self.assertEqual(payment.json()["data"]["active_task"]["position_code"], "supply.business_reviewer")
        self.assertEqual(business.json()["data"]["active_task"]["position_code"], "supply.business_reviewer")
        payment_tasks = list(self.db.scalars(select(WorkflowTask).where(
            WorkflowTask.instance_id == payment.json()["data"]["workflow_instance_id"]
        ).order_by(WorkflowTask.sequence)))
        business_tasks = list(self.db.scalars(select(WorkflowTask).where(
            WorkflowTask.instance_id == business.json()["data"]["workflow_instance_id"]
        ).order_by(WorkflowTask.sequence)))
        self.assertEqual(len(payment_tasks), 7)
        self.assertEqual(len(business_tasks), 5)
        self.assertIn("investment.duty.supply_finance_review", {
            task.required_position_code for task in payment_tasks
        })
        self.assertNotIn("investment.duty.supply_finance_review", {
            task.required_position_code for task in business_tasks
        })
        self.assertNotIn("external.legal_counsel", {
            task.required_position_code for task in payment_tasks + business_tasks
        })

    def test_todo_and_compatibility_actions_use_active_workflow_task(self):
        submitted = self.submit(self.business)
        self.assertEqual(submitted.status_code, 200, submitted.text)
        self.current_user = self.reviewer_a
        first_todo = self.client.get("/api/v1/approval-forms/todo")
        self.current_user = self.reviewer_b
        second_todo = self.client.get("/api/v1/approval-forms/todo")
        self.assertEqual([item["id"] for item in first_todo.json()["data"]], [self.business.id])
        self.assertEqual([item["id"] for item in second_todo.json()["data"]], [self.business.id])

        approved = self.client.post(
            f"/api/v1/approval-forms/{self.business.id}/approve",
            json={"comment": "复核同意"},
        )
        self.assertEqual(approved.status_code, 200, approved.text)
        self.assertEqual(approved.json()["data"]["active_task"]["position_code"], "supply.company_leader")
        self.current_user = self.other_leader
        self.assertEqual(self.client.get("/api/v1/approval-forms/todo").json()["data"], [])
        self.current_user = self.leader
        self.assertEqual(
            [item["id"] for item in self.client.get("/api/v1/approval-forms/todo").json()["data"]],
            [self.business.id],
        )
        returned = self.client.post(
            f"/api/v1/approval-forms/{self.business.id}/reject",
            json={"comment": "退回复核"},
        )
        self.assertEqual(returned.status_code, 200, returned.text)
        self.assertEqual(returned.json()["data"]["active_task"]["position_code"], "supply.business_reviewer")

    def test_return_to_handler_resubmits_same_instance_and_preserves_projection(self):
        first = self.submit(self.business)
        instance_id = first.json()["data"]["workflow_instance_id"]
        self.current_user = self.reviewer_a
        returned = self.client.post(
            f"/api/v1/approval-forms/{self.business.id}/reject",
            json={"comment": "请补充"},
        )
        self.assertEqual(returned.status_code, 200, returned.text)
        self.current_user = self.handler
        resumed = self.client.post(
            f"/api/v1/approval-forms/{self.business.id}/submit",
            json={},
        )

        self.assertEqual(resumed.status_code, 200, resumed.text)
        self.assertEqual(resumed.json()["data"]["workflow_instance_id"], instance_id)
        actions = list(self.db.scalars(
            select(WorkflowTaskAction)
            .join(WorkflowTaskAction.task)
            .where(WorkflowTask.instance_id == instance_id)
            .order_by(WorkflowTaskAction.id)
        ))
        self.assertEqual(
            [action.action for action in actions],
            [WorkflowAction.SUBMIT, WorkflowAction.RETURN, WorkflowAction.SUBMIT],
        )
        projections = list(self.db.scalars(
            select(ApprovalFormAction)
            .where(ApprovalFormAction.form_id == self.business.id)
            .order_by(ApprovalFormAction.id)
        ))
        self.assertEqual([item.action.value for item in projections], ["approve", "reject", "approve"])
        self.assertTrue(all(item.workflow_task_action_id is not None for item in projections))
        self.assertEqual(projections[-1].position_code, "supply.business_handler")

    def test_generic_actions_project_snapshots_but_reassign_does_not_fake_approval(self):
        submitted = self.submit(self.business)
        task_id = submitted.json()["data"]["active_task"]["id"]
        submit_projection = self.db.scalar(select(ApprovalFormAction).where(
            ApprovalFormAction.form_id == self.business.id
        ))
        self.assertEqual(submit_projection.position_code, "supply.business_handler")
        self.current_user = self.reviewer_a
        approved = self.client.post(
            f"/api/v1/workflows/tasks/{task_id}/approve",
            json={"comment": "通用入口通过"},
        )
        self.assertEqual(approved.status_code, 200, approved.text)
        projection = self.db.scalar(
            select(ApprovalFormAction)
            .where(ApprovalFormAction.form_id == self.business.id)
            .order_by(ApprovalFormAction.id.desc())
        )
        self.assertEqual(projection.action.value, "approve")
        self.assertEqual(projection.position_code, "supply.business_reviewer")
        self.assertEqual(projection.position_name, "业务复核")

        leader_task = self.active_task(self.business)
        before = self.db.query(ApprovalFormAction).count()
        self.current_user = self.admin
        reassigned = self.client.post(
            f"/api/v1/workflows/tasks/{leader_task.id}/reassign",
            json={"user_id": self.other_leader.id, "reason": "负责人调整"},
        )
        self.assertEqual(reassigned.status_code, 200, reassigned.text)
        self.assertEqual(self.db.query(ApprovalFormAction).count(), before)
        self.current_user = self.other_leader
        returned = self.client.post(
            f"/api/v1/workflows/tasks/{leader_task.id}/reject",
            json={"reason": "通用入口退回"},
        )
        self.assertEqual(returned.status_code, 200, returned.text)
        return_projection = self.db.scalar(
            select(ApprovalFormAction)
            .where(ApprovalFormAction.form_id == self.business.id)
            .order_by(ApprovalFormAction.id.desc())
        )
        self.assertEqual(return_projection.action.value, "reject")
        self.assertEqual(return_projection.position_code, "supply.company_leader")

    def test_legacy_reject_wrapper_requires_nonblank_reason(self):
        submitted = self.submit(self.business)
        self.assertEqual(submitted.status_code, 200, submitted.text)
        self.current_user = self.reviewer_a

        response = self.client.post(
            f"/api/v1/approval-forms/{self.business.id}/reject",
            json={"comment": "   "},
        )

        self.assertEqual(response.status_code, 422)

    def test_pending_count_uses_actionable_shared_and_designated_tasks(self):
        submitted = self.submit(self.business)
        self.assertEqual(submitted.status_code, 200, submitted.text)

        for reviewer in (self.reviewer_a, self.reviewer_b):
            with self.subTest(reviewer=reviewer.username):
                self.current_user = reviewer
                data = self.client.get("/api/v1/approval/pending-count").json()["data"]
                self.assertEqual(data, {"contract": 0, "business": 1, "total": 1})

        self.current_user = self.reviewer_a
        self.client.post(
            f"/api/v1/approval-forms/{self.business.id}/approve",
            json={"comment": "复核完成"},
        )
        self.current_user = self.leader
        selected = self.client.get("/api/v1/approval/pending-count").json()["data"]
        self.current_user = self.other_leader
        unselected = self.client.get("/api/v1/approval/pending-count").json()["data"]

        self.assertEqual(selected["business"], 1)
        self.assertEqual(unselected["business"], 0)
        leader_task = self.active_task(self.business)
        leader_task.status = WorkflowTaskStatus.AWAITING_REASSIGNMENT
        self.db.commit()
        self.current_user = self.leader
        self.assertEqual(
            self.client.get("/api/v1/approval/pending-count").json()["data"]["business"],
            0,
        )
        self.current_user = self.admin
        admin_data = self.client.get("/api/v1/approval/pending-count").json()["data"]
        self.assertEqual(admin_data["total"], 0)
        self.assertEqual(admin_data["reassignment"], 1)

    def test_pending_count_counts_contract_workflow_task_separately(self):
        contract = Contract(
            contract_no="WF-FORM-STATS-CONTRACT",
            title="Stats Contract",
            status=ContractStatus.DRAFT,
            created_by=self.handler.id,
        )
        self.db.add(contract)
        self.db.commit()
        start_workflow(
            self.db,
            "contract",
            contract.id,
            self.handler,
            {
                "company_leader": self.leader.id,
                "legal_counsel": self.legal.id,
                "supply_governance_leader": self.governance.id,
            },
        )
        self.db.commit()

        self.current_user = self.leader
        selected = self.client.get("/api/v1/approval/pending-count").json()["data"]
        self.current_user = self.other_leader
        unselected = self.client.get("/api/v1/approval/pending-count").json()["data"]

        self.assertEqual(selected["contract"], 1)
        self.assertEqual(unselected["contract"], 0)

    def test_print_endpoint_passes_position_snapshot(self):
        submitted = self.submit(self.business)
        self.assertEqual(submitted.status_code, 200, submitted.text)

        with patch(
            "app.api.v1.endpoints.approval.print_svc.build_approval_form_xlsx",
            return_value=b"xlsx",
        ) as build:
            response = self.client.get(f"/api/v1/approval-forms/{self.business.id}/print")

        self.assertEqual(response.status_code, 200, response.text)
        steps = build.call_args.args[1]
        self.assertEqual(steps[0]["position_code"], "supply.business_handler")
        self.assertEqual(steps[0]["role"], "supply.business_handler")
        self.assertFalse(build.call_args.args[0]["legacy_workflow"])

        with patch(
            "app.api.v1.endpoints.approval.print_svc.build_approval_form_xlsx",
            return_value=b"xlsx",
        ) as legacy_build:
            legacy_response = self.client.get(
                f"/api/v1/approval-forms/{self.payment.id}/print"
            )
        self.assertEqual(legacy_response.status_code, 200, legacy_response.text)
        self.assertTrue(legacy_build.call_args.args[0]["legacy_workflow"])

        instance = self.db.get(WorkflowInstance, self.business.workflow_instance_id)
        instance.workflow_version.version = 1
        self.db.commit()
        with patch(
            "app.api.v1.endpoints.approval.print_svc.build_approval_form_xlsx",
            return_value=b"xlsx",
        ) as v1_build:
            v1_response = self.client.get(
                f"/api/v1/approval-forms/{self.business.id}/print"
            )
        self.assertEqual(v1_response.status_code, 200, v1_response.text)
        self.assertTrue(v1_build.call_args.args[0]["legacy_workflow"])


class ApprovalPrintWorkflowTest(unittest.TestCase):
    def _workbook(self, form_type, steps, *, legacy_workflow=False):
        data = build_approval_form_xlsx(
            {
                "form_type": form_type,
                "department": "供管公司",
                "business_desc": "打印测试",
                "legacy_workflow": legacy_workflow,
            },
            steps,
        )
        return openpyxl.load_workbook(BytesIO(data)).active

    def test_new_workflow_prints_by_position_code_and_updates_governance_label(self):
        ws = self._workbook(ContractType.BUSINESS, [{
            "step": 0,
            "position_code": "governance.supply_leader",
            "role": "business_handler",
            "name": "分管领导",
            "comment": "同意",
            "signature": "",
        }])

        self.assertEqual(ws["B9"].value, "供管公司分管领导")
        self.assertIn("同意", ws["C9"].value)
        self.assertIn("分管领导", ws["C9"].value)
        self.assertIsNone(ws["C6"].value)

    def test_historical_print_falls_back_to_legacy_role(self):
        ws = self._workbook(ContractType.PAYMENT, [{
            "step": 0,
            "position_code": None,
            "role": "invest_director",
            "name": "历史领导",
            "comment": "历史意见",
            "signature": "",
        }], legacy_workflow=True)

        self.assertEqual(ws["B12"].value, "供管公司分管领导")
        self.assertIn("历史意见", ws["C12"].value)
        self.assertIn("历史领导", ws["C12"].value)
        self.assertIsNone(ws["C9"].value)

    def test_v2_missing_position_snapshot_does_not_fall_back_to_legacy_role(self):
        ws = self._workbook(ContractType.BUSINESS, [{
            "step": 0,
            "position_code": None,
            "role": "invest_director",
            "name": "异常旧角色",
            "comment": "不应打印",
            "signature": "",
        }])

        self.assertEqual(ws["B9"].value, "供管公司分管领导")
        self.assertIsNone(ws["C9"].value)
        self.assertIsNone(ws["C6"].value)


if __name__ == "__main__":
    unittest.main()
