import unittest
from datetime import date, datetime, timedelta
from io import BytesIO

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import create_engine, select, update
from sqlalchemy.orm import Session

from app.api.v1.endpoints.legal_risk import _import_batch_for_user
from app.core.enums import AssignmentStatus, Role
from app.db.base import Base
from app.models.legal_risk import (
    LegalAlertDelivery,
    LegalAlertType,
    LegalCase,
    LegalCaseAlert,
    LegalCaseImportBatch,
    LegalCaseImportRow,
    LegalImportStatus,
)
from app.models.organization import Organization, Position, UserAssignment
from app.models.user import User
from app.services.legal_imports import (
    build_import_template,
    confirm_import,
    expire_unconfirmed_batches,
    preview_import,
)
from app.services.organization_catalog import seed_authorization_catalog


class LegalImportTest(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine("sqlite+pysqlite:///:memory:")
        Base.metadata.create_all(self.engine)
        self.db = Session(self.engine)
        seed_authorization_catalog(self.db)
        self.user = User(username="admin", full_name="管理员", hashed_password="x",
                         role=Role.INFO_MAINTAINER, is_superuser=True, is_active=True)
        self.db.add(self.user); self.db.commit()

    def tearDown(self):
        self.db.close(); self.engine.dispose()

    def _filled_template(self, responsible_user_id: int | None = None) -> bytes:
        workbook = load_workbook(build_import_template())
        workbook["案件基本信息"].append([
            "EXT-1", "导入测试案件", "审查立案", "合同纠纷", "测试法院", "(2026)测1号",
            1000, responsible_user_id or self.user.id, "案情", "请求",
        ])
        workbook["当事人"].append(["EXT-1", "原告", "原告公司", "organization", "", "", ""])
        workbook["当事人"].append(["EXT-1", "被告", "被告公司", "organization", "", "", ""])
        buffer = BytesIO(); workbook.save(buffer)
        return buffer.getvalue()

    def _assign_legal_business_position(self, user: User) -> None:
        self.db.add(UserAssignment(
            user_id=user.id,
            organization_id=self.db.scalar(
                select(Organization.id).where(Organization.code == "investment.legal_risk")
            ),
            position_id=self.db.scalar(
                select(Position.id).where(Position.code == "investment.department.junior_manager")
            ),
            valid_from=date(2026, 1, 1),
            status=AssignmentStatus.ACTIVE,
        ))

    def test_template_contains_eight_versioned_sheets(self):
        workbook = load_workbook(build_import_template())
        self.assertEqual(workbook.sheetnames, [
            "案件基本信息", "当事人", "裁判结果", "查扣冻资产", "清回止损",
            "进展风险", "期限事件", "填写说明与枚举值",
        ])
        self.assertEqual(workbook["填写说明与枚举值"]["B1"].value, "legal-case-v1")
        headers = [cell.value for cell in workbook["裁判结果"][1]]
        self.assertIn("判决金额", headers)
        self.assertNotIn("可执行金额", headers)

    def test_preview_and_confirm_are_transactional_and_idempotent(self):
        batch = preview_import(self.db, self._filled_template(), "cases.xlsx", self.user)
        self.db.commit()
        self.assertEqual(batch.error_rows, 0)
        warning_ids = self.db.scalars(
            self.db.query(LegalCaseImportRow.id).filter(
                LegalCaseImportRow.batch_id == batch.id,
                LegalCaseImportRow.validation_status == "warning",
            ).statement
        ).all()
        result = confirm_import(self.db, batch, self.user, list(warning_ids))
        self.db.commit()
        self.assertEqual(result["imported_cases"], 1)
        self.assertEqual(self.db.query(LegalCase).count(), 1)
        with self.assertRaises(HTTPException) as raised:
            confirm_import(self.db, batch, self.user, list(warning_ids))
        self.assertEqual(raised.exception.status_code, 409)

    def test_imported_future_custom_deadline_creates_task_without_delivery(self):
        workbook = load_workbook(BytesIO(self._filled_template()))
        workbook["期限事件"].append([
            "EXT-1", "custom", "custom deadline", date(2026, 12, 31), 7, None,
        ])
        buffer = BytesIO(); workbook.save(buffer)
        batch = preview_import(self.db, buffer.getvalue(), "deadline.xlsx", self.user)
        self.db.commit()
        warning_ids = self.db.scalars(
            self.db.query(LegalCaseImportRow.id).filter(
                LegalCaseImportRow.batch_id == batch.id,
                LegalCaseImportRow.validation_status == "warning",
            ).statement
        ).all()

        confirm_import(self.db, batch, self.user, list(warning_ids))
        self.db.commit()

        alert = self.db.query(LegalCaseAlert).filter_by(source_type="deadline").one()
        self.assertEqual(alert.alert_type, LegalAlertType.CUSTOM)
        self.assertEqual(self.db.query(LegalAlertDelivery).count(), 0)

    def test_preview_rejects_formal_case_without_both_party_sides(self):
        workbook = load_workbook(build_import_template())
        workbook["案件基本信息"].append([
            "EXT-2", "缺少被告案件", "审查立案", "合同纠纷", "测试法院", "(2026)测2号",
            1000, self.user.id, "案情", "请求",
        ])
        workbook["当事人"].append(["EXT-2", "原告", "原告公司", "organization", "", "", ""])
        buffer = BytesIO(); workbook.save(buffer)

        batch = preview_import(self.db, buffer.getvalue(), "missing-party.xlsx", self.user)

        self.assertGreater(batch.error_rows, 0)
        basic = self.db.query(LegalCaseImportRow).filter(
            LegalCaseImportRow.batch_id == batch.id,
            LegalCaseImportRow.sheet_name == "案件基本信息",
        ).one()
        self.assertIn("正式案件至少需要一名被告/被申请人", basic.errors)

    def test_preview_rejects_wrong_template_version(self):
        workbook = load_workbook(build_import_template())
        workbook["填写说明与枚举值"]["B1"] = "legal-case-v0"
        buffer = BytesIO(); workbook.save(buffer)

        with self.assertRaises(HTTPException) as raised:
            preview_import(self.db, buffer.getvalue(), "old-template.xlsx", self.user)

        self.assertEqual(raised.exception.status_code, 422)
        self.assertIn("模板版本不匹配", raised.exception.detail)

    def test_preview_validates_required_dates_user_and_reminder_range(self):
        workbook = load_workbook(BytesIO(self._filled_template()))
        workbook["清回止损"].append(["EXT-1", "回款", "", 0, "测试回款"])
        workbook["期限事件"].append(["EXT-1", "开庭", "测试开庭", "", 366, 999999])
        buffer = BytesIO(); workbook.save(buffer)

        batch = preview_import(self.db, buffer.getvalue(), "invalid-details.xlsx", self.user)
        rows = self.db.query(LegalCaseImportRow).filter(
            LegalCaseImportRow.batch_id == batch.id,
            LegalCaseImportRow.validation_status == "error",
        ).all()
        errors = [message for row in rows for message in row.errors]

        self.assertIn("日期不能为空", errors)
        self.assertIn("金额必须大于 0", errors)
        self.assertIn("事件日期不能为空", errors)
        self.assertIn("提前天数必须在 0 到 365 之间", errors)
        self.assertIn("责任人用户ID不是有效的投资公司在职用户", errors)

    def test_import_batches_are_private_to_creator_except_superuser(self):
        owner = User(
            username="owner", full_name="导入人", hashed_password="x",
            role=Role.BUSINESS_HANDLER, is_superuser=False, is_active=True,
        )
        outsider = User(
            username="outsider", full_name="其他导入人", hashed_password="x",
            role=Role.BUSINESS_HANDLER, is_superuser=False, is_active=True,
        )
        self.db.add_all([owner, outsider]); self.db.flush()
        for user in (owner, outsider):
            self._assign_legal_business_position(user)
        self.db.flush()
        batch = preview_import(self.db, self._filled_template(), "private.xlsx", owner)

        self.assertIs(_import_batch_for_user(self.db, batch.id, owner), batch)
        self.assertIs(_import_batch_for_user(self.db, batch.id, self.user), batch)
        with self.assertRaises(HTTPException) as raised:
            _import_batch_for_user(self.db, batch.id, outsider)
        self.assertEqual(raised.exception.status_code, 403)

    def test_normalized_legal_position_is_valid_import_responsible_user(self):
        responsible = User(
            username="responsible",
            full_name="法务责任人",
            hashed_password="x",
            role=Role.UNASSIGNED,
            is_active=True,
        )
        self.db.add(responsible)
        self.db.flush()
        self._assign_legal_business_position(responsible)
        self.db.commit()

        batch = preview_import(
            self.db,
            self._filled_template(responsible.id),
            "normalized-responsible.xlsx",
            self.user,
        )

        self.assertEqual(batch.error_rows, 0)

    def test_import_claim_prevents_stale_confirmation(self):
        batch = preview_import(self.db, self._filled_template(), "claimed.xlsx", self.user)
        self.db.flush()
        self.db.execute(
            update(LegalCaseImportBatch)
            .where(LegalCaseImportBatch.id == batch.id)
            .values(status=LegalImportStatus.IMPORTING)
        )
        self.db.commit()

        with self.assertRaises(HTTPException) as raised:
            confirm_import(self.db, batch, self.user, [])

        self.assertEqual(raised.exception.status_code, 409)
        self.assertEqual(self.db.query(LegalCase).count(), 0)

    def test_cleanup_physically_deletes_old_unconfirmed_batches_and_rows(self):
        batch = preview_import(self.db, self._filled_template(), "expired.xlsx", self.user)
        now = datetime(2026, 8, 14, 9, 0, 0)
        batch.created_at = now - timedelta(days=8)
        batch_id = batch.id
        self.db.commit()

        self.assertEqual(expire_unconfirmed_batches(self.db, now), 1)
        self.db.commit()

        self.assertIsNone(self.db.get(LegalCaseImportBatch, batch_id))
        self.assertEqual(
            self.db.query(LegalCaseImportRow).filter(
                LegalCaseImportRow.batch_id == batch_id
            ).count(),
            0,
        )


if __name__ == "__main__":
    unittest.main()
