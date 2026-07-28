"""财务对账单处理服务。

职责：
1. 解析多 Sheet 的平台对账单 xlsx（抖音/美团/携程），保留独立的历史财务数据源。
   并从明细 Sheet 统计订单数与间夜。
2. UPSERT 写入 biz_financial_metrics（按平台+账期覆盖，幂等）。
3. 从门票/酒店台账汇总经营核心指标、资金占用时长和逐期服务费图表。
"""
from __future__ import annotations

import re
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

import openpyxl
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.financial import FinanceConfig, FinancialMetrics
from app.models.hotel_ledger import HotelLedger
from app.models.project import ProjectMetrics
from app.models.ticket_ledger import TicketLedger

PLATFORM_KEYWORDS = {
    "douyin": "抖音",
    "meituan": "美团",
    "ctrip": "携程",
}
PLATFORM_LABELS = {"douyin": "抖音", "meituan": "美团", "ctrip": "携程"}


def _num(v):
    """宽松地把单元格转为 Decimal（去除 ¥、逗号、空格等）；无法解析返回 None。"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            return Decimal(str(v))
        except InvalidOperation:
            return None
    s = re.sub(r"[^\d.\-]", "", str(v).strip())
    if not s or s in ("-", "."):
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _detect_platform(sheet_name: str):
    for key, kw in PLATFORM_KEYWORDS.items():
        if kw in sheet_name:
            return key
    return None


def _parse_period(title: str) -> str:
    """从对账单标题解析账期，如 '2026年6月24日-6月30日...' → '2026-06-24~06-30'。"""
    m = re.search(
        r"(\d{4})\D*?(\d{1,2})月(\d{1,2})日\s*[-~至到]\s*(?:(\d{1,2})月)?(\d{1,2})日",
        str(title or ""),
    )
    if not m:
        return ""
    y, m1, d1, m2, d2 = m.group(1), m.group(2), m.group(3), m.group(4) or m.group(2), m.group(5)
    return f"{y}-{int(m1):02d}-{int(d1):02d}~{int(m2):02d}-{int(d2):02d}"


def _parse_reconciliation(ws) -> dict | None:
    """从对账单 Sheet 提取 realized_scale / gross_income / period。

    对账单末尾有『总计』行，其数值列尾部形如 [..., 出版应得, 应扣预付, 应扣预付]
    （应扣预付在合并单元格中重复一次）。据此：
      gross_income  = 最后一个数值（应扣出版预付/回款）
      realized_scale= 倒数第三个数值（出版应得到账金额）
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None

    title = ""
    for c in rows[0]:
        if c is not None and str(c).strip():
            title = str(c)
            break
    period = _parse_period(title)

    total_row = None
    for r in rows:
        for c in r:
            if c is not None and str(c).strip() in ("总计", "合计"):
                total_row = r
                break
        if total_row is not None:
            break
    if total_row is None:
        return None

    nums = [x for x in (_num(c) for c in total_row) if x is not None]
    if len(nums) < 3:
        return None
    gross = nums[-1]
    realized = nums[-3]
    return {"realized_scale": realized, "gross_income": gross, "period": period}


def _parse_detail(ws) -> dict:
    """从明细 Sheet 统计 order_count / room_nights / gmv(可选)。"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"order_count": 0, "room_nights": 0, "gmv": None}

    header = [str(c).strip() if c is not None else "" for c in rows[0]]

    def find_col(*keywords):
        for i, h in enumerate(header):
            if all(k in h for k in keywords):
                return i
        return None

    ci_night = find_col("间夜")
    ci_gmv = find_col("订单实收")  # 抖音明细提供 GMV

    order_count = 0
    nights = Decimal("0")
    gmv = Decimal("0")
    has_gmv = ci_gmv is not None
    for r in rows[1:]:
        night = _num(r[ci_night]) if (ci_night is not None and ci_night < len(r)) else None
        if night is None:
            continue  # 跳过空行/小计行
        order_count += 1
        nights += night
        if has_gmv and ci_gmv < len(r):
            g = _num(r[ci_gmv])
            if g:
                gmv += g
    return {
        "order_count": order_count,
        "room_nights": int(nights),
        "gmv": (gmv if has_gmv else None),
    }


def parse_workbook(content: bytes) -> list[dict]:
    """解析对账单 xlsx，返回每平台一条指标 dict。"""
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    recon_ws, detail_ws = {}, {}
    for name in wb.sheetnames:
        plat = _detect_platform(name)
        if not plat:
            continue
        if "对账单" in name:
            recon_ws[plat] = wb[name]
        else:
            detail_ws[plat] = wb[name]

    results = []
    for plat, ws in recon_ws.items():
        base = _parse_reconciliation(ws)
        if not base:
            continue
        det = _parse_detail(detail_ws[plat]) if plat in detail_ws else {"order_count": 0, "room_nights": 0, "gmv": None}
        results.append({
            "platform": plat,
            "period": base["period"] or PLATFORM_LABELS.get(plat, plat),
            "realized_scale": base["realized_scale"],
            "gross_income": base["gross_income"],
            "gmv": det["gmv"],
            "order_count": det["order_count"],
            "room_nights": det["room_nights"],
        })
    return results


def upsert_metrics(db: Session, parsed: list[dict]) -> list[FinancialMetrics]:
    """按 (平台, 账期) UPSERT，覆盖式、幂等。"""
    saved = []
    for item in parsed:
        row = db.scalar(
            select(FinancialMetrics).where(
                FinancialMetrics.platform == item["platform"],
                FinancialMetrics.period == item["period"],
            )
        )
        if not row:
            row = FinancialMetrics(platform=item["platform"], period=item["period"])
            db.add(row)
        row.realized_scale = item["realized_scale"]
        row.gross_income = item["gross_income"]
        row.gmv = item["gmv"]
        row.order_count = item["order_count"]
        row.room_nights = item["room_nights"]
        saved.append(row)
    db.commit()
    for r in saved:
        db.refresh(r)
    return saved


def get_or_create_config(db: Session) -> FinanceConfig:
    cfg = db.scalar(select(FinanceConfig).limit(1))
    if not cfg:
        cfg = FinanceConfig(total_invested_cost=Decimal("0"))
        db.add(cfg)
        db.commit()
        db.refresh(cfg)
    return cfg


def _to_platform_metric(row: FinancialMetrics) -> dict:
    return {
        "platform": row.platform,
        "platform_label": PLATFORM_LABELS.get(row.platform, row.platform),
        "period": row.period,
        "realized_scale": row.realized_scale,
        "gross_income": row.gross_income,
        "gmv": row.gmv,
        "order_count": row.order_count,
        "room_nights": row.room_nights,
    }


def _to_project_metric(row: ProjectMetrics) -> dict:
    return {
        "seq": row.seq,
        "project_name": row.project_name,
        "platforms": row.platforms,
        "invested_amount": row.invested_amount,
        "realized_scale": row.realized_scale,
        "gross_profit": row.gross_profit,
        "profit_rate": row.profit_rate,
        "pay_date": row.pay_date.isoformat() if row.pay_date else None,
        "term_months": row.term_months,
        "capital_occupied": (row.invested_amount or Decimal("0")) - (row.realized_scale or Decimal("0")),
    }


_LEDGER_DATE_RE = re.compile(r"(20\d{2})\D{0,3}(\d{1,2})(?:\D{0,3}(\d{1,2}))?")


def _decimal(value) -> Decimal:
    return value if isinstance(value, Decimal) else Decimal(str(value or 0))


def _period_date(row) -> date | None:
    """取台账归属期日期，缺少结构化日期时再从期次文本中兜底解析。"""
    for value in (getattr(row, "period_start", None), getattr(row, "period_end", None)):
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
    text = getattr(row, "period_text", "") or getattr(row, "check_date_text", "") or ""
    match = _LEDGER_DATE_RE.search(text)
    if match:
        try:
            return date(int(match.group(1)), int(match.group(2)), int(match.group(3) or 1))
        except ValueError:
            pass
    created = getattr(row, "created_at", None)
    if isinstance(created, datetime):
        return created.date()
    return created if isinstance(created, date) else None


def _period_label(row) -> str:
    label = getattr(row, "period_text", "") or getattr(row, "check_date_text", "")
    if label:
        return label
    start = getattr(row, "period_start", None)
    end = getattr(row, "period_end", None)
    if start or end:
        return f"{start or ''}~{end or ''}".strip("~")
    return f"台账#{getattr(row, 'id', '')}"


def _profit_point(scenic_id: str, business_type: str, row, service_fee: Decimal) -> dict:
    period_date = _period_date(row)
    period = _period_label(row)
    date_key = period_date.isoformat() if period_date else "undated"
    return {
        "scenic_id": scenic_id,
        "business_type": business_type,
        "period": period,
        "period_key": f"{date_key}::{period}",
        "year": period_date.year if period_date else None,
        "month": period_date.month if period_date else None,
        "service_fee": service_fee,
    }


def _hotel_period_key(row) -> tuple[str, str]:
    period = (
        getattr(row, "source_file", "")
        or getattr(row, "detail_name", "")
        or getattr(row, "period_text", "")
        or getattr(row, "check_date_text", "")
        or f"row:{getattr(row, 'id', '')}"
    )
    return row.scenic_id, period


def build_ledger_metrics(
    ticket_rows: list[TicketLedger],
    hotel_rows: list[HotelLedger],
    *,
    today: date | None = None,
) -> dict:
    """按台账快照计算经营指标；不回写台账，也不重新执行业务计算算法。"""
    today = today or date.today()
    total_invested = Decimal("0")
    total_realized = Decimal("0")
    total_gross = Decimal("0")
    total_repaid = Decimal("0")
    occupation_weight = Decimal("0")
    occupation_amount = Decimal("0")
    ledger_profit: list[dict] = []

    def add_occupation(net_investment: Decimal, start: date | None, end: date | None) -> None:
        nonlocal occupation_weight, occupation_amount
        if net_investment <= 0 or not start:
            return
        finish = end or today
        days = max((finish - start).days, 0)
        occupation_weight += net_investment * Decimal(days)
        occupation_amount += net_investment

    for row in ticket_rows:
        payment = _decimal(row.payment_amount)
        co_investment = _decimal(getattr(row, "co_investment_amount", 0))
        net_investment = payment - co_investment
        total_invested += net_investment
        total_realized += _decimal(row.jinying_amount)
        total_gross += _decimal(row.service_fee)
        total_repaid += _decimal(row.repay_amount)
        add_occupation(net_investment, row.pay_date, row.repay_date)
        ledger_profit.append(
            _profit_point(row.scenic_id, "ticket", row, _decimal(row.service_fee))
        )

    hotel_groups: dict[tuple[str, str], list[HotelLedger]] = defaultdict(list)
    for row in hotel_rows:
        hotel_groups[_hotel_period_key(row)].append(row)
        total_realized += _decimal(row.jinying_amount)
        total_gross += _decimal(row.service_fee)

    for (scenic_id, _), rows in hotel_groups.items():
        # 同期共享字段只计算一次；取付款金额最大的代表行兼容迁移前可能存在的同期不一致数据。
        representative = max(rows, key=lambda item: _decimal(item.payment_amount))
        payment = _decimal(representative.payment_amount)
        co_investment = _decimal(getattr(representative, "co_investment_amount", 0))
        net_investment = payment - co_investment
        total_invested += net_investment
        payment_date = next((r.payment_date for r in rows if r.payment_date), None)
        repay_date = next((r.repay_date for r in rows if r.repay_date), None)
        repay_amount = next((r.repay_amount for r in rows if r.repay_amount is not None), None)
        total_repaid += _decimal(repay_amount)
        add_occupation(net_investment, payment_date, repay_date)
        total_fee = sum((_decimal(r.service_fee) for r in rows), Decimal("0"))
        ledger_profit.append(_profit_point(scenic_id, "hotel", representative, total_fee))

    ledger_profit.sort(
        key=lambda item: (
            item["year"] is None,
            item["year"] or 9999,
            item["month"] or 99,
            item["period_key"],
            item["scenic_id"],
            item["business_type"],
        )
    )
    years = sorted({item["year"] for item in ledger_profit if item["year"] is not None}, reverse=True)
    scenic_ids = sorted({row.scenic_id for row in [*ticket_rows, *hotel_rows]})
    occupation_days = (
        round(float(occupation_weight / occupation_amount), 1) if occupation_amount else None
    )
    profit_rate = (
        round(float(total_gross / total_invested * 100), 2) if total_invested > 0 else None
    )
    return {
        "existing_scale": total_invested,
        "total_realized_scale": total_realized,
        "total_gross_income": total_gross,
        "profit_rate": profit_rate,
        "capital_occupation_days": occupation_days,
        "capital_occupied": total_invested - total_repaid,
        "ledger_profit": ledger_profit,
        "available_years": years,
        "scenic_ids": scenic_ids,
    }


def build_dashboard(db: Session) -> dict:
    """构建财务经营看板：核心指标由门票/酒店台账驱动。"""
    ticket_rows = db.scalars(
        select(TicketLedger).order_by(TicketLedger.period_start, TicketLedger.id)
    ).all()
    hotel_rows = db.scalars(
        select(HotelLedger).order_by(HotelLedger.period_start, HotelLedger.id)
    ).all()
    ledger_metrics = build_ledger_metrics(ticket_rows, hotel_rows)

    # 项目统计表和平台对账单继续作为独立数据源返回，兼容现有上传与大屏调用。
    projects = db.scalars(select(ProjectMetrics).order_by(ProjectMetrics.seq)).all()
    cfg = get_or_create_config(db)
    fm_rows = db.scalars(select(FinancialMetrics).order_by(FinancialMetrics.platform)).all()
    platforms = [_to_platform_metric(r) for r in fm_rows]

    return {
        **ledger_metrics,
        "available_funds": cfg.available_funds or Decimal("0"),
        "projects": [_to_project_metric(p) for p in projects],
        "invested_cost": cfg.total_invested_cost or Decimal("0"),
        "platforms": platforms,
    }
