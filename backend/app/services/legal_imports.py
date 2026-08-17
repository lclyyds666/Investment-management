"""法务案件标准 Excel 模板、预检和事务导入。"""
from __future__ import annotations

import hashlib
from copy import copy
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, select, update
from sqlalchemy.orm import Session

from app.core.enums import CompanyCode
from app.models.legal_risk import (
    LegalCase,
    LegalCaseAsset,
    LegalCaseDeadline,
    LegalCaseImportBatch,
    LegalCaseImportRow,
    LegalCaseJudgment,
    LegalCaseParty,
    LegalCaseProgress,
    LegalCaseRecovery,
    LegalCaseStage,
    LegalCaseStatus,
    LegalDeadlineType,
    LegalImportStatus,
    LegalJudgmentType,
    LegalPartyType,
    LegalProgressType,
    LegalRecoveryType,
)
from app.models.user import User
from app.services.legal_cases import next_case_no, record_activity, set_current_enforcement_basis
from app.services.legal_clock import legal_now
from app.services.permissions import get_company_role

TEMPLATE_VERSION = "legal-case-v1"
SHEET_HEADERS = {
    "案件基本信息": ["外部案件编号", "案件名称", "主状态", "案由", "受理法院", "法院案号", "标的额", "负责人用户ID", "案情简介", "诉讼请求"],
    "当事人": ["外部案件编号", "当事人类型", "名称", "身份类型", "证件或统一社会信用代码", "联系方式", "地址"],
    "裁判结果": ["外部案件编号", "类型", "摘要", "裁判或达成日期", "生效日期", "履行期限", "可执行金额", "当前执行依据"],
    "查扣冻资产": ["外部案件编号", "资产类型", "资产名称或位置", "措施类型", "顺位", "开始日期", "到期日期", "提前天数", "处置状态", "说明"],
    "清回止损": ["外部案件编号", "记录类型", "日期", "金额", "来源说明"],
    "进展风险": ["外部案件编号", "记录类型", "内容", "风险点", "下一步计划", "责任人用户ID", "计划完成日"],
    "期限事件": ["外部案件编号", "事件类型", "事项名称", "事件日期", "提前天数", "责任人用户ID"],
}
SHEET_NAMES = tuple(SHEET_HEADERS)

STATUS_VALUES = {
    "审查立案": LegalCaseStatus.REVIEW_FILING,
    "审理中": LegalCaseStatus.IN_TRIAL,
    "已判决": LegalCaseStatus.JUDGED,
    "执行中": LegalCaseStatus.ENFORCEMENT,
    "终本": LegalCaseStatus.TERMINAL,
    "已结案": LegalCaseStatus.CLOSED,
    **{item.value: item for item in LegalCaseStatus},
}
PARTY_VALUES = {"原告": LegalPartyType.PLAINTIFF, "被告": LegalPartyType.DEFENDANT, "第三人": LegalPartyType.THIRD_PARTY,
                **{item.value: item for item in LegalPartyType}}
JUDGMENT_VALUES = {"一审": LegalJudgmentType.FIRST_INSTANCE, "二审": LegalJudgmentType.SECOND_INSTANCE,
                   "再审": LegalJudgmentType.RETRIAL, "调解": LegalJudgmentType.MEDIATION,
                   "和解": LegalJudgmentType.SETTLEMENT, **{item.value: item for item in LegalJudgmentType}}
RECOVERY_VALUES = {"回款": LegalRecoveryType.RECOVERY, "避免损失": LegalRecoveryType.AVOIDED_LOSS,
                   **{item.value: item for item in LegalRecoveryType}}
PROGRESS_VALUES = {"进展": LegalProgressType.PROGRESS, "法律意见": LegalProgressType.LEGAL_OPINION,
                   **{item.value: item for item in LegalProgressType}}
DEADLINE_VALUES = {"开庭": LegalDeadlineType.HEARING, "缴费/材料": LegalDeadlineType.PAYMENT_MATERIAL,
                   "自定义": LegalDeadlineType.CUSTOM, **{item.value: item for item in LegalDeadlineType}}


def build_import_template() -> BytesIO:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, headers in SHEET_HEADERS.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{chr(64 + min(len(headers), 26))}1"
        for cell in sheet[1]:
            font = copy(cell.font)
            font.bold = True
            cell.font = font
    instructions = workbook.create_sheet("填写说明与枚举值")
    instructions.append(["模板版本", TEMPLATE_VERSION])
    instructions.append(["关联规则", "所有子表通过外部案件编号关联案件基本信息；附件请在案件导入后上传。"])
    instructions.append(["主状态", "审查立案、审理中、已判决、执行中、终本、已结案"])
    instructions.append(["裁判类型", "一审、二审、再审、调解、和解"])
    instructions.append(["当事人类型", "原告、被告、第三人"])
    instructions.append(["记录类型", "回款、避免损失；进展、法律意见"])
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def normalize_text(value) -> str:
    return " ".join(str(value or "").strip().split())


def _json_value(value):
    if isinstance(value, (date, datetime)): return value.isoformat()
    if isinstance(value, Decimal): return str(value)
    return value


def _date(value, label: str, errors: list[str]) -> str | None:
    if value in (None, ""): return None
    if isinstance(value, datetime): return value.date().isoformat()
    if isinstance(value, date): return value.isoformat()
    text = normalize_text(value)
    try:
        return date.fromisoformat(text).isoformat()
    except ValueError:
        errors.append(f"{label}必须为日期")
        return None


def _decimal(value, label: str, errors: list[str], *, required: bool = False) -> str | None:
    if value in (None, ""):
        if required: errors.append(f"{label}不能为空")
        return None
    try:
        parsed = Decimal(str(value).replace(",", ""))
        if parsed < 0: raise InvalidOperation
        return str(parsed.quantize(Decimal("0.01")))
    except (InvalidOperation, ValueError):
        errors.append(f"{label}必须为非负金额")
        return None


def _integer(value, label: str, errors: list[str]) -> int | None:
    if value in (None, ""): return None
    try:
        return int(value)
    except (TypeError, ValueError):
        errors.append(f"{label}必须为整数")
        return None


def _reminder_days(value, label: str, errors: list[str]) -> int | None:
    parsed = _integer(value, label, errors)
    if parsed is not None and not 0 <= parsed <= 365:
        errors.append(f"{label}必须在 0 到 365 之间")
        return None
    return parsed


def _validate_investment_user(db: Session, user_id: int | None, label: str, errors: list[str]) -> None:
    if user_id is None:
        return
    user = db.scalar(select(User).where(User.id == user_id, User.is_active.is_(True)))
    if user is None or (
        not user.is_superuser
        and get_company_role(db, user, CompanyCode.INVESTMENT) is None
    ):
        errors.append(f"{label}不是有效的投资公司在职用户")


def _enum(value, mapping, label: str, errors: list[str]) -> str | None:
    text = normalize_text(value)
    item = mapping.get(text)
    if item is None:
        errors.append(f"{label}枚举值无效")
        return None
    return item.value


def _row_dict(headers: list[str], values: tuple) -> dict:
    return {header: _json_value(values[index] if index < len(values) else None) for index, header in enumerate(headers)}


def _validate_row(sheet_name: str, data: dict, known_keys: set[str], db: Session) -> tuple[dict, list[str], list[str]]:
    warnings: list[str] = []
    errors: list[str] = []
    external_id = normalize_text(data.get("外部案件编号"))
    normalized = {"external_id": external_id}
    if not external_id: errors.append("外部案件编号不能为空")
    elif sheet_name != "案件基本信息" and external_id not in known_keys:
        errors.append("外部案件编号在案件基本信息中不存在")

    if sheet_name == "案件基本信息":
        normalized.update({
            "case_name": normalize_text(data.get("案件名称")),
            "status": _enum(data.get("主状态"), STATUS_VALUES, "主状态", errors),
            "cause_of_action": normalize_text(data.get("案由")),
            "court": normalize_text(data.get("受理法院")),
            "court_case_no": normalize_text(data.get("法院案号")),
            "subject_amount": _decimal(data.get("标的额"), "标的额", errors, required=True),
            "responsible_user_id": _integer(data.get("负责人用户ID"), "负责人用户ID", errors),
            "case_summary": normalize_text(data.get("案情简介")),
            "claims": normalize_text(data.get("诉讼请求")),
        })
        for key, label in (("case_name", "案件名称"), ("cause_of_action", "案由"),
                           ("court", "受理法院"), ("court_case_no", "法院案号")):
            if not normalized[key]: errors.append(f"{label}不能为空")
        if normalized["responsible_user_id"] is None: errors.append("负责人用户ID不能为空")
        _validate_investment_user(db, normalized["responsible_user_id"], "负责人用户ID", errors)
        if normalized["court_case_no"] and db.scalar(select(LegalCase.id).where(
            LegalCase.court_case_no == normalized["court_case_no"], LegalCase.deleted_at.is_(None)
        )):
            warnings.append("法院案号已存在，确认后仍可导入")
    elif sheet_name == "当事人":
        normalized.update({
            "party_type": _enum(data.get("当事人类型"), PARTY_VALUES, "当事人类型", errors),
            "name": normalize_text(data.get("名称")), "identity_type": normalize_text(data.get("身份类型")) or "organization",
            "identity_no": normalize_text(data.get("证件或统一社会信用代码")),
            "contact": normalize_text(data.get("联系方式")), "address": normalize_text(data.get("地址")),
        })
        if not normalized["name"]: errors.append("当事人名称不能为空")
    elif sheet_name == "裁判结果":
        normalized.update({
            "judgment_type": _enum(data.get("类型"), JUDGMENT_VALUES, "裁判类型", errors),
            "summary": normalize_text(data.get("摘要")),
            "judgment_date": _date(data.get("裁判或达成日期"), "裁判或达成日期", errors),
            "effective_date": _date(data.get("生效日期"), "生效日期", errors),
            "performance_deadline": _date(data.get("履行期限"), "履行期限", errors),
            "executable_amount": _decimal(data.get("可执行金额"), "可执行金额", errors),
            "is_current_enforcement_basis": normalize_text(data.get("当前执行依据")).lower() in {"是", "true", "1", "yes"},
        })
    elif sheet_name == "查扣冻资产":
        normalized.update({
            "asset_type": normalize_text(data.get("资产类型")), "asset_name": normalize_text(data.get("资产名称或位置")),
            "measure_type": normalize_text(data.get("措施类型")), "priority_type": normalize_text(data.get("顺位")),
            "start_date": _date(data.get("开始日期"), "开始日期", errors),
            "expiry_date": _date(data.get("到期日期"), "到期日期", errors),
            "reminder_days": _reminder_days(data.get("提前天数"), "提前天数", errors),
            "disposal_status": normalize_text(data.get("处置状态")), "notes": normalize_text(data.get("说明")),
        })
        for key, label in (("asset_type", "资产类型"), ("asset_name", "资产名称或位置"), ("measure_type", "措施类型")):
            if not normalized[key]: errors.append(f"{label}不能为空")
    elif sheet_name == "清回止损":
        normalized.update({
            "recovery_type": _enum(data.get("记录类型"), RECOVERY_VALUES, "记录类型", errors),
            "recovery_date": _date(data.get("日期"), "日期", errors),
            "amount": _decimal(data.get("金额"), "金额", errors, required=True),
            "source_description": normalize_text(data.get("来源说明")),
        })
        if normalized["recovery_date"] is None: errors.append("日期不能为空")
        if normalized["amount"] is not None and Decimal(normalized["amount"]) <= 0:
            errors.append("金额必须大于 0")
    elif sheet_name == "进展风险":
        normalized.update({
            "progress_type": _enum(data.get("记录类型"), PROGRESS_VALUES, "记录类型", errors),
            "content": normalize_text(data.get("内容")), "risk_points": normalize_text(data.get("风险点")),
            "next_plan": normalize_text(data.get("下一步计划")),
            "responsible_user_id": _integer(data.get("责任人用户ID"), "责任人用户ID", errors),
            "planned_date": _date(data.get("计划完成日"), "计划完成日", errors),
        })
        if not normalized["content"]: errors.append("内容不能为空")
        _validate_investment_user(db, normalized["responsible_user_id"], "责任人用户ID", errors)
    elif sheet_name == "期限事件":
        normalized.update({
            "deadline_type": _enum(data.get("事件类型"), DEADLINE_VALUES, "事件类型", errors),
            "title": normalize_text(data.get("事项名称")),
            "event_date": _date(data.get("事件日期"), "事件日期", errors),
            "reminder_days": _reminder_days(data.get("提前天数"), "提前天数", errors),
            "responsible_user_id": _integer(data.get("责任人用户ID"), "责任人用户ID", errors),
        })
        if not normalized["title"]: errors.append("事项名称不能为空")
        if normalized["event_date"] is None: errors.append("事件日期不能为空")
        _validate_investment_user(db, normalized["responsible_user_id"], "责任人用户ID", errors)
    return normalized, warnings, errors


def preview_import(db: Session, content: bytes, file_name: str, actor: User) -> LegalCaseImportBatch:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail="无法读取 Excel 文件") from exc
    missing = [name for name in (*SHEET_NAMES, "填写说明与枚举值") if name not in workbook.sheetnames]
    if missing: raise HTTPException(status_code=422, detail=f"模板缺少工作表：{'、'.join(missing)}")
    instructions = workbook["填写说明与枚举值"]
    if normalize_text(instructions["A1"].value) != "模板版本" or normalize_text(instructions["B1"].value) != TEMPLATE_VERSION:
        raise HTTPException(status_code=422, detail=f"模板版本不匹配，请下载 {TEMPLATE_VERSION} 标准模板")

    basic_sheet = workbook["案件基本信息"]
    basic_headers = [normalize_text(cell.value) for cell in basic_sheet[1]]
    absent_basic = [header for header in SHEET_HEADERS["案件基本信息"] if header not in basic_headers]
    if absent_basic:
        raise HTTPException(status_code=422, detail=f"案件基本信息缺少列：{'、'.join(absent_basic)}")
    known_keys = {
        normalize_text(values[basic_headers.index("外部案件编号")])
        for values in basic_sheet.iter_rows(min_row=2, values_only=True)
        if any(value not in (None, "") for value in values)
    }
    batch = LegalCaseImportBatch(
        file_name=Path(file_name).name if file_name else "法务案件导入.xlsx",
        file_hash=hashlib.sha256(content).hexdigest(), template_version=TEMPLATE_VERSION,
        status=LegalImportStatus.PREVIEWED, created_by=actor.id,
    )
    db.add(batch); db.flush()
    rows: list[LegalCaseImportRow] = []
    for sheet_name, expected_headers in SHEET_HEADERS.items():
        sheet = workbook[sheet_name]
        headers = [normalize_text(cell.value) for cell in sheet[1]]
        absent = [header for header in expected_headers if header not in headers]
        if absent: raise HTTPException(status_code=422, detail=f"{sheet_name}缺少列：{'、'.join(absent)}")
        for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value not in (None, "") for value in values): continue
            raw = _row_dict(headers, values)
            normalized, warnings, errors = _validate_row(sheet_name, raw, known_keys, db)
            status = "error" if errors else "warning" if warnings else "valid"
            rows.append(LegalCaseImportRow(
                batch_id=batch.id, sheet_name=sheet_name, row_number=row_number,
                normalized_data=normalized, validation_status=status,
                warnings=warnings, errors=errors,
            ))

    basic_rows: dict[str, list[LegalCaseImportRow]] = {}
    party_types: dict[str, set[str]] = {}
    for row in rows:
        external_id = row.normalized_data.get("external_id", "")
        if row.sheet_name == "案件基本信息":
            basic_rows.setdefault(external_id, []).append(row)
        elif row.sheet_name == "当事人" and row.normalized_data.get("party_type"):
            party_types.setdefault(external_id, set()).add(row.normalized_data["party_type"])
    for external_id, case_rows in basic_rows.items():
        for row in case_rows:
            if len(case_rows) > 1:
                row.errors.append("外部案件编号在案件基本信息中重复")
            types = party_types.get(external_id, set())
            if LegalPartyType.PLAINTIFF.value not in types:
                row.errors.append("正式案件至少需要一名原告/申请人")
            if LegalPartyType.DEFENDANT.value not in types:
                row.errors.append("正式案件至少需要一名被告/被申请人")
            if row.errors:
                row.validation_status = "error"
    db.add_all(rows); db.flush()
    batch.total_rows = len(rows)
    batch.importable_rows = sum(row.validation_status != "error" for row in rows)
    batch.warning_rows = sum(row.validation_status == "warning" for row in rows)
    batch.error_rows = sum(row.validation_status == "error" for row in rows)
    return batch


def _as_date(value: str | None):
    return date.fromisoformat(value) if value else None


def confirm_import(
    db: Session,
    batch: LegalCaseImportBatch,
    actor: User,
    confirmed_warning_rows: list[int],
) -> dict:
    claimed = db.execute(
        update(LegalCaseImportBatch)
        .where(
            LegalCaseImportBatch.id == batch.id,
            LegalCaseImportBatch.status == LegalImportStatus.PREVIEWED,
        )
        .values(status=LegalImportStatus.IMPORTING)
    )
    if claimed.rowcount != 1:
        raise HTTPException(status_code=409, detail="该批次不能重复确认")
    rows = db.scalars(select(LegalCaseImportRow).where(
        LegalCaseImportRow.batch_id == batch.id
    ).order_by(LegalCaseImportRow.id.asc())).all()
    if any(row.validation_status == "error" for row in rows):
        raise HTTPException(status_code=409, detail="批次仍有错误，不能导入")
    warning_ids = {row.id for row in rows if row.validation_status == "warning"}
    if not warning_ids.issubset(set(confirmed_warning_rows)):
        raise HTTPException(status_code=409, detail="请先确认全部警告行")
    grouped: dict[str, list[LegalCaseImportRow]] = {}
    for row in rows: grouped.setdefault(row.normalized_data["external_id"], []).append(row)
    imported = 0
    for external_id, case_rows in grouped.items():
        basic_row = next((row for row in case_rows if row.sheet_name == "案件基本信息"), None)
        if basic_row is None: continue
        data = basic_row.normalized_data
        case = LegalCase(
            stage=LegalCaseStage.FORMAL, case_no=next_case_no(db, legal_now().year),
            case_name=data["case_name"], status=LegalCaseStatus(data["status"]),
            cause_of_action=data["cause_of_action"], court=data["court"],
            court_case_no=data["court_case_no"], subject_amount=Decimal(data["subject_amount"]),
            responsible_user_id=data["responsible_user_id"], case_summary=data["case_summary"],
            claims=data["claims"], created_by=actor.id, activated_by=actor.id, activated_at=legal_now(),
        )
        db.add(case); db.flush()
        current_basis_id = None
        for row in case_rows:
            item = row.normalized_data
            if row.sheet_name == "当事人":
                db.add(LegalCaseParty(case_id=case.id, party_type=LegalPartyType(item["party_type"]), name=item["name"],
                                      identity_type=item["identity_type"], identity_no=item["identity_no"],
                                      contact=item["contact"], address=item["address"]))
            elif row.sheet_name == "裁判结果":
                child = LegalCaseJudgment(case_id=case.id, judgment_type=LegalJudgmentType(item["judgment_type"]),
                    summary=item["summary"], judgment_date=_as_date(item["judgment_date"]),
                    effective_date=_as_date(item["effective_date"]), performance_deadline=_as_date(item["performance_deadline"]),
                    executable_amount=Decimal(item["executable_amount"]) if item["executable_amount"] else None,
                    is_current_enforcement_basis=item["is_current_enforcement_basis"])
                db.add(child); db.flush()
                if child.is_current_enforcement_basis: current_basis_id = child.id
            elif row.sheet_name == "查扣冻资产":
                db.add(LegalCaseAsset(case_id=case.id, asset_type=item["asset_type"], asset_name=item["asset_name"],
                    measure_type=item["measure_type"], priority_type=item["priority_type"],
                    start_date=_as_date(item["start_date"]), expiry_date=_as_date(item["expiry_date"]),
                    reminder_days=item["reminder_days"], disposal_status=item["disposal_status"], notes=item["notes"]))
            elif row.sheet_name == "清回止损":
                db.add(LegalCaseRecovery(case_id=case.id, recovery_type=LegalRecoveryType(item["recovery_type"]),
                    recovery_date=_as_date(item["recovery_date"]), amount=Decimal(item["amount"]),
                    source_description=item["source_description"], registered_by=actor.id))
            elif row.sheet_name == "进展风险":
                db.add(LegalCaseProgress(case_id=case.id, progress_type=LegalProgressType(item["progress_type"]),
                    content=item["content"], risk_points=item["risk_points"], next_plan=item["next_plan"],
                    responsible_user_id=item["responsible_user_id"], planned_date=_as_date(item["planned_date"]), registered_by=actor.id))
            elif row.sheet_name == "期限事件":
                db.add(LegalCaseDeadline(case_id=case.id, deadline_type=LegalDeadlineType(item["deadline_type"]),
                    title=item["title"], event_date=_as_date(item["event_date"]), reminder_days=item["reminder_days"],
                    responsible_user_id=item["responsible_user_id"]))
            row.imported_case_id = case.id
        if current_basis_id is not None: set_current_enforcement_basis(db, case.id, current_basis_id)
        record_activity(db, case.id, "import", actor, summary=f"导入批次 {batch.id}，外部编号 {external_id}")
        imported += 1
    batch.status = LegalImportStatus.IMPORTED
    batch.confirmed_by = actor.id
    batch.confirmed_at = legal_now()
    db.flush()
    return {"batch_id": batch.id, "imported_cases": imported}


def build_error_report(rows: list[LegalCaseImportRow]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "校验结果"
    sheet.append(["工作表", "Excel行号", "状态", "警告", "错误"])
    for row in rows:
        sheet.append([row.sheet_name, row.row_number, row.validation_status,
                      "；".join(row.warnings), "；".join(row.errors)])
    buffer = BytesIO(); workbook.save(buffer); buffer.seek(0)
    return buffer


def expire_unconfirmed_batches(db: Session, now: datetime | None = None) -> int:
    cutoff = (now or legal_now()) - timedelta(days=7)
    batch_ids = db.scalars(select(LegalCaseImportBatch.id).where(
        LegalCaseImportBatch.status == LegalImportStatus.PREVIEWED,
        LegalCaseImportBatch.created_at < cutoff,
    )).all()
    if not batch_ids:
        return 0
    db.execute(delete(LegalCaseImportRow).where(LegalCaseImportRow.batch_id.in_(batch_ids)))
    db.execute(delete(LegalCaseImportBatch).where(LegalCaseImportBatch.id.in_(batch_ids)))
    db.flush()
    return len(batch_ids)
