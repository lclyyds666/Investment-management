"""文旅业务·景区酒店平台核销台账计算服务（泉州欧乐堡）。

解析一份对账明细 xlsx（含抖音/美团/携程多平台、每平台多张周明细表），
按平台聚合出「结算基数 / 间夜 / 服务商到账·佣金(抖音)」，供前端确认后生成台账。

平台计算口径（已用真实业务台账 6 期反推验证）：
- 抖音(71列, 与门票同构; 列名跨期漂移, 按别名匹配)：
    服务商到账 = −Σ(服务商服务费|服务商佣金) = Σ(订单实收−软件−达人−团长)
    服务商佣金(默认) = 订单实收×6% − 达人服务费 − 团长服务费 (达人/团长为负,相加=减)
    结算基数(出版应得) = 到账 − 佣金
- 美团(26列)：结算基数 = Σ结算金额
- 携程(24列)：结算基数 = Σ结算价
统一：景区核销 = 结算基数 × 核销率(0.9)。服务费/结算两种算法（compute_row.fee_algo）：
  算法1(默认)：服务费 = 间夜 × 44；结算金额 = 核销 + 服务费。
  算法2       ：结算金额 = 结算基数 × 结算费率(0.94)；服务费 = 结算金额 − 核销。
"""
from __future__ import annotations

import json
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO

import openpyxl

DEFAULT_RATE_HEXIAO = Decimal("0.90")     # 景区核销率
DEFAULT_FEE_PER_NIGHT = Decimal("44.00")  # 每间夜服务费(算法1)
DEFAULT_RATE_SETTLE = Decimal("0.94")     # 结算费率(算法2:结算金额=结算基数×结算费率)
DEFAULT_COMMISSION_RATE = Decimal("0.06")  # 抖音服务商佣金率(对订单实收)
_CENT = Decimal("0.01")

# 平台识别（按 sheet 标题前缀）
PLATFORMS = ("抖音", "美团", "携程")

# 列别名（跨期/跨平台容错）
COL_DY_SHISHOU = {"订单实收金额"}
COL_DY_RUANJIAN = {"软件服务费"}
COL_DY_DAREN = {"达人服务费", "达人佣金"}
COL_DY_TUANZHANG = {"团长服务费", "撮合经纪服务费"}
COL_DY_FUWUSHANG = {"服务商服务费", "服务商佣金"}
COL_DY_NIGHTS = {"间夜", "间夜数"}
COL_DY_TIME = {"核销时间"}
COL_MT_JIESUAN = {"结算金额"}
COL_MT_NIGHTS = {"间夜"}
COL_MT_CHECKIN = {"入住日期"}
COL_MT_LEAVE = {"离店日期"}
COL_XC_JIESUAN = {"结算价"}
COL_XC_NIGHTS = {"间夜"}
COL_XC_CHECKIN = {"入住日期"}
COL_XC_LEAVE = {"离店日期"}


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            return Decimal(str(v))
        except InvalidOperation:
            return None
    s = re.sub(r"[^\d.\-]", "", str(v).strip())
    if not s or s in ("-", ".", "-."):
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _q(v: Decimal) -> Decimal:
    return (v or Decimal("0")).quantize(_CENT, rounding=ROUND_HALF_UP)


def _to_date(v):
    if v is None or str(v).strip() == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    m = re.search(r"(\d{4})\D+(\d{1,2})\D+(\d{1,2})", s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    return None


def _idx(header: list, names: set[str]) -> int:
    for i, h in enumerate(header):
        if h is not None and str(h).strip() in names:
            return i
    return -1


def _platform_of(title: str) -> str | None:
    t = title or ""
    for p in PLATFORMS:
        if t.startswith(p) or p in t[:4]:
            return p
    return None


def _year_from_filename(filename: str) -> int | None:
    m = re.search(r"(20\d{2})", filename or "")
    return int(m.group(1)) if m else None


def _dates_from_title(title: str, year: int | None) -> tuple[date | None, date | None]:
    """从 sheet 标题解析日期跨度，如 '携程6.15-6.21' / '抖音明细1.14-1.20'。"""
    if not year:
        return None, None
    md = re.findall(r"(\d{1,2})[.\-=](\d{1,2})", title or "")
    ds = []
    for mo, dy in md:
        try:
            ds.append(date(year, int(mo), int(dy)))
        except ValueError:
            continue
    if not ds:
        return None, None
    return min(ds), max(ds)


def _header_row(rows_iter) -> tuple[list, list]:
    """返回 (header, remaining_rows_list)。header=首个非空行。"""
    rows = list(rows_iter)
    hi = 0
    for i, r in enumerate(rows):
        if r and any(c is not None and str(c).strip() for c in r):
            hi = i
            break
    header = [("" if c is None else str(c).strip()) for c in rows[hi]]
    return header, rows[hi + 1:]


def _col_sum(rows: list, i: int) -> Decimal:
    if i < 0:
        return Decimal("0")
    s = Decimal("0")
    for r in rows:
        v = _num(r[i]) if i < len(r) else None
        if v is not None:
            s += v
    return s


def _count_rows(rows: list, key_idx: int) -> int:
    n = 0
    for r in rows:
        if key_idx < 0:
            if r and any(c is not None and str(c).strip() for c in r):
                n += 1
        elif key_idx < len(r) and r[key_idx] is not None and str(r[key_idx]).strip():
            n += 1
    return n


def _new_day() -> dict:
    return {"recv": Decimal("0"), "base": Decimal("0"), "shishou": Decimal("0"),
            "daren": Decimal("0"), "tuanzhang": Decimal("0"), "nights": 0}


# --------------------------------------------------------------------------- #
# 逐日明细：序列化持久化 + 逐日重算（编辑改费率/佣金/算法时仍按天累加）
# --------------------------------------------------------------------------- #
def _days_from_daily(daily: dict[str, dict]) -> list[dict]:
    return [{
        "recv": dd["recv"], "base": dd["base"], "shishou": dd["shishou"],
        "daren": dd["daren"], "tuanzhang": dd["tuanzhang"], "nights": int(dd["nights"] or 0),
    } for dd in daily.values()]


def serialize_daily(daily: dict[str, dict]) -> str:
    out = [{
        "r": str(dd["recv"]), "b": str(dd["base"]), "s": str(dd["shishou"]),
        "d": str(dd["daren"]), "t": str(dd["tuanzhang"]), "n": int(dd["nights"] or 0),
    } for dd in daily.values()]
    return json.dumps(out, ensure_ascii=False)


def _days_from_json(daily_json: str) -> list[dict]:
    if not daily_json:
        return []
    try:
        raw = json.loads(daily_json)
    except (ValueError, TypeError):
        return []
    days = []
    for d in raw:
        days.append({
            "recv": _num(d.get("r")) or Decimal("0"),
            "base": _num(d.get("b")) or Decimal("0"),
            "shishou": _num(d.get("s")) or Decimal("0"),
            "daren": _num(d.get("d")) or Decimal("0"),
            "tuanzhang": _num(d.get("t")) or Decimal("0"),
            "nights": int(d.get("n") or 0),
        })
    return days


def recompute_from_days(platform: str, days: list[dict],
                        rate_hexiao: Decimal = DEFAULT_RATE_HEXIAO,
                        rate_settle: Decimal = DEFAULT_RATE_SETTLE,
                        fee_per_night: Decimal = DEFAULT_FEE_PER_NIGHT,
                        fee_algo: int = 1,
                        commission_override=None,
                        room_nights_override=None) -> dict | None:
    """酒店逐日重算并累加（逐日舍入到分）。结算基数=到账/毛额−佣金；核销=Σ(基数_day×核销率)；
    算法1：服务费=间夜×每间夜服务费(间夜用行聚合值，×整数费率无逐日舍入差)、结算=核销+服务费；
    算法2：结算=Σ(基数_day×结算费率)、服务费=结算−核销。
    佣金（仅抖音）默认逐日自动=实收×6%−达人−团长；改总额时按各天实收占比分摊差额。"""
    if not days:
        return None
    is_dy = platform == "抖音"
    if is_dy:
        comm_auto = [_q(d["shishou"] * DEFAULT_COMMISSION_RATE + d["daren"] + d["tuanzhang"]) for d in days]
        c0 = _q(sum(comm_auto, Decimal("0")))
        if commission_override is None or abs((commission_override or Decimal("0")) - c0) < Decimal("0.005"):
            comm_day, c_total = comm_auto, c0
        else:
            delta = commission_override - c0
            s_total = sum((d["shishou"] for d in days), Decimal("0"))
            n = len(days) or 1
            comm_day = []
            for i, d in enumerate(days):
                share = (d["shishou"] / s_total) if s_total > 0 else (Decimal("1") / n)
                comm_day.append(_q(comm_auto[i] + delta * share))
            c_total = _q(commission_override)
    else:
        comm_day = [Decimal("0")] * len(days)
        c_total = Decimal("0")

    algo2 = int(fee_algo or 1) == 2
    hexiao = jinying_a2 = fee_a2 = sbase = Decimal("0")
    nights_sum = 0
    for i, d in enumerate(days):
        raw_base = d["recv"] if is_dy else d["base"]
        b = raw_base - comm_day[i]
        sbase += b
        nights_sum += int(d["nights"] or 0)
        hx = _q(b * rate_hexiao)
        hexiao += hx
        if algo2:
            jy = _q(b * (rate_settle or Decimal("0")))
            jinying_a2 += jy
            fee_a2 += _q(jy - hx)
    if algo2:
        jinying, fee_sum = _q(jinying_a2), _q(fee_a2)
    else:
        nights = room_nights_override if room_nights_override is not None else nights_sum
        fee_sum = _q(Decimal(int(nights or 0)) * (fee_per_night or Decimal("0")))
        jinying = _q(hexiao + fee_sum)
    return {
        "supplier_commission": _q(c_total),
        "settle_base": _q(sbase),
        "hexiao_amount": _q(hexiao),
        "service_fee": _q(fee_sum),
        "jinying_amount": _q(jinying),
    }


def recompute_from_json(platform: str, daily_json: str,
                        rate_hexiao: Decimal = DEFAULT_RATE_HEXIAO,
                        rate_settle: Decimal = DEFAULT_RATE_SETTLE,
                        fee_per_night: Decimal = DEFAULT_FEE_PER_NIGHT,
                        fee_algo: int = 1,
                        commission_override=None,
                        room_nights_override=None) -> dict | None:
    return recompute_from_days(platform, _days_from_json(daily_json), rate_hexiao,
                               rate_settle, fee_per_night, fee_algo, commission_override,
                               room_nights_override)


def daily_defaults(platform: str, daily: dict[str, dict],
                   rate_hexiao: Decimal = DEFAULT_RATE_HEXIAO,
                   fee_per_night: Decimal = DEFAULT_FEE_PER_NIGHT) -> dict:
    """解析时的按日精准默认值（算法1；佣金取逐日自动值）。"""
    res = recompute_from_days(platform, _days_from_daily(daily), rate_hexiao,
                              DEFAULT_RATE_SETTLE, fee_per_night, 1, None)
    if res is None:
        return {"commission": Decimal("0"), "hexiao": Decimal("0"),
                "service_fee": Decimal("0"), "jinying": Decimal("0")}
    return {"commission": res["supplier_commission"], "hexiao": res["hexiao_amount"],
            "service_fee": res["service_fee"], "jinying": res["jinying_amount"]}


def parse_hotel_file(content: bytes, filename: str = "") -> dict:
    """解析一份酒店对账明细，按平台聚合。**按日期分组 → 逐日计算舍入 → 累加**。"""
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
    year = _year_from_filename(filename)
    agg: dict[str, dict] = {}
    warnings: list[str] = []

    try:
        for ws in wb.worksheets:
            plat = _platform_of(ws.title)
            if plat is None:
                continue
            header, rows = _header_row(ws.iter_rows(values_only=True))
            d = agg.setdefault(plat, {
                "platform": plat, "room_nights": 0, "order_count": 0,
                "base_received": Decimal("0"), "daily": {}, "pstart": None, "pend": None,
            })
            s0, e0 = _dates_from_title(ws.title, year)
            daily: dict = d["daily"]

            if plat == "抖音":
                i_fs = _idx(header, COL_DY_FUWUSHANG)
                i_shi = _idx(header, COL_DY_SHISHOU)
                i_ruan = _idx(header, COL_DY_RUANJIAN)
                i_da = _idx(header, COL_DY_DAREN)
                i_tu = _idx(header, COL_DY_TUANZHANG)
                i_night = _idx(header, COL_DY_NIGHTS)
                i_time = _idx(header, COL_DY_TIME)
                for r in rows:
                    shi = _num(r[i_shi]) if 0 <= i_shi < len(r) else None
                    ruan = _num(r[i_ruan]) if 0 <= i_ruan < len(r) else None
                    da = _num(r[i_da]) if 0 <= i_da < len(r) else None
                    tu = _num(r[i_tu]) if 0 <= i_tu < len(r) else None
                    fs = _num(r[i_fs]) if 0 <= i_fs < len(r) else None
                    if shi is None and fs is None:
                        continue
                    dt = _to_date(r[i_time]) if 0 <= i_time < len(r) else None
                    key = dt.isoformat() if dt else "NA"
                    day = daily.setdefault(key, _new_day())
                    # 当日到账 = −服务商服务费(若有) 否则 实收+软件+达人+团长(费用为负)
                    if fs is not None:
                        recv = -fs
                    else:
                        recv = (shi or Decimal("0")) + (ruan or Decimal("0")) + (da or Decimal("0")) + (tu or Decimal("0"))
                    day["recv"] += recv
                    day["shishou"] += (shi or Decimal("0"))
                    day["daren"] += (da or Decimal("0"))
                    day["tuanzhang"] += (tu or Decimal("0"))
                    n = _num(r[i_night]) if 0 <= i_night < len(r) else None
                    day["nights"] += int(n) if n is not None else 0
                    d["base_received"] += recv
                    d["order_count"] += 1
                    d["room_nights"] += int(n) if n is not None else 0
                    if dt:
                        s0 = dt if s0 is None else min(s0, dt)
                        e0 = dt if e0 is None else max(e0, dt)
            else:  # 美团 / 携程
                if plat == "美团":
                    i_j = _idx(header, COL_MT_JIESUAN); i_night = _idx(header, COL_MT_NIGHTS)
                    i_leave = _idx(header, COL_MT_LEAVE); i_in = _idx(header, COL_MT_CHECKIN)
                else:
                    i_j = _idx(header, COL_XC_JIESUAN); i_night = _idx(header, COL_XC_NIGHTS)
                    i_leave = _idx(header, COL_XC_LEAVE); i_in = _idx(header, COL_XC_CHECKIN)
                for r in rows:
                    base = _num(r[i_j]) if 0 <= i_j < len(r) else None
                    if base is None:
                        continue
                    dt = None
                    if 0 <= i_leave < len(r):
                        dt = _to_date(r[i_leave])
                    if dt is None and 0 <= i_in < len(r):
                        dt = _to_date(r[i_in])
                    key = dt.isoformat() if dt else "NA"
                    day = daily.setdefault(key, _new_day())
                    day["base"] += base
                    n = _num(r[i_night]) if 0 <= i_night < len(r) else None
                    day["nights"] += int(n) if n is not None else 0
                    d["base_received"] += base
                    d["order_count"] += 1
                    d["room_nights"] += int(n) if n is not None else 0
                    if dt:
                        s0 = dt if s0 is None else min(s0, dt)
                        e0 = dt if e0 is None else max(e0, dt)

            if s0:
                d["pstart"] = s0 if d["pstart"] is None else min(d["pstart"], s0)
            if e0:
                d["pend"] = e0 if d["pend"] is None else max(d["pend"], e0)
    finally:
        wb.close()

    out = []
    for plat in PLATFORMS:
        if plat not in agg:
            continue
        d = agg[plat]
        base_received = _q(d["base_received"])
        defs = daily_defaults(plat, d["daily"])   # 按日计算的精准默认值
        p_text = ""
        if d["pstart"] and d["pend"]:
            s, e = d["pstart"], d["pend"]
            p_text = f"{s.year}/{s.month}/{s.day}-{e.year}/{e.month}/{e.day}"
        if d["order_count"] == 0:
            warnings.append(f"{plat}：未解析到有效明细")
        out.append({
            "platform": plat,
            "room_nights": d["room_nights"],
            "order_count": d["order_count"],
            "base_received": base_received,
            "suggested_commission": defs["commission"],
            "def_hexiao": defs["hexiao"],
            "def_service_fee": defs["service_fee"],
            "def_jinying": defs["jinying"],
            "daily_json": serialize_daily(d["daily"]),
            "period_start": d["pstart"],
            "period_end": d["pend"],
            "period_text": p_text,
            "check_date_text": p_text,
        })
    return {"platforms": out, "warnings": warnings}


def compute_row(
    platform: str,
    base_received: Decimal,
    supplier_commission: Decimal,
    room_nights: int,
    rate_hexiao: Decimal = DEFAULT_RATE_HEXIAO,
    fee_per_night: Decimal = DEFAULT_FEE_PER_NIGHT,
    fee_algo: int = 1,
    rate_settle: Decimal = DEFAULT_RATE_SETTLE,
) -> dict:
    """由平台原值/佣金/间夜与比例计算台账列。

    结算基数 settle_base = 服务商到账 − 服务商佣金（美团/携程佣金恒 0）。
    景区核销 = 结算基数 × 核销率。两种服务费/结算算法：
      算法1(默认)：服务费 = 间夜 × 每间夜服务费；结算金额 = 景区核销 + 服务费。
      算法2       ：结算金额 = 结算基数 × 结算费率；服务费 = 结算金额 − 景区核销。
    """
    recv = base_received or Decimal("0")
    comm = supplier_commission or Decimal("0")
    if platform != "抖音":
        comm = Decimal("0")  # 美团/携程无服务商佣金层
    settle_base = recv - comm
    hexiao = _q(settle_base * rate_hexiao)
    if int(fee_algo or 1) == 2:
        jinying = _q(settle_base * (rate_settle or Decimal("0")))
        fee = _q(jinying - hexiao)
    else:
        fee = _q(Decimal(int(room_nights or 0)) * (fee_per_night or Decimal("0")))
        jinying = _q(hexiao + fee)
    return {
        "supplier_commission": _q(comm),
        "settle_base": _q(settle_base),
        "hexiao_amount": hexiao,
        "service_fee": fee,
        "jinying_amount": jinying,
    }


def running_pending(prev_balance: Decimal, payment_amount: Decimal, hexiao_amount: Decimal) -> Decimal:
    """期次递推：本期待核销 = 上期余额 + 本期付款 − 本期核销（首期 prev=0）。按平台各自滚动。"""
    prev = prev_balance or Decimal("0")
    pay = payment_amount or Decimal("0")
    hexiao = hexiao_amount or Decimal("0")
    return _q(prev + pay - hexiao)


# —— 导出（台账列不含付款金额；含合计）——
_EXPORT_HEADERS = [
    "平台", "酒店名称", "核对日期",
    "景区核销金额", "景区待核销金额", "结算金额", "服务费", "间夜",
    "回款日期", "回款金额",
]


def _fmt_date(v) -> str:
    if not v:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v)


def _fa(v):
    d = _num(v)
    return float(_q(d)) if d is not None else ""


def build_export_workbook(rows: list[dict], title: str = "酒店平台业务台账") -> bytes:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "酒店平台业务台账"
    ncol = len(_EXPORT_HEADERS)
    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    tc = ws.cell(row=1, column=1, value=title)
    tc.font = Font(size=13, bold=True); tc.alignment = center

    head_fill = PatternFill("solid", fgColor="DCE6F1")
    for c, name in enumerate(_EXPORT_HEADERS, start=1):
        cell = ws.cell(row=2, column=c, value=name)
        cell.font = Font(bold=True); cell.alignment = center; cell.fill = head_fill; cell.border = border

    s_hx = s_jy = s_fee = Decimal("0"); s_night = 0; s_repay = Decimal("0")
    last_pending = Decimal("0")
    seen_periods: set[str] = set()  # 回款每期共享 → 合计每期只计一次
    r = 3
    for row in rows:
        vals = [
            row.get("platform", ""), row.get("hotel_name", ""), row.get("check_date_text", ""),
            _fa(row.get("hexiao_amount")), _fa(row.get("pending_writeoff")),
            _fa(row.get("jinying_amount")), _fa(row.get("service_fee")),
            int(row.get("room_nights") or 0),
            _fmt_date(row.get("repay_date")), _fa(row.get("repay_amount")),
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v); cell.alignment = center; cell.border = border
        s_hx += _num(row.get("hexiao_amount")) or Decimal("0")
        s_jy += _num(row.get("jinying_amount")) or Decimal("0")
        s_fee += _num(row.get("service_fee")) or Decimal("0")
        s_night += int(row.get("room_nights") or 0)
        pk = row.get("period_key") or ""
        if pk not in seen_periods:
            seen_periods.add(pk)
            s_repay += _num(row.get("repay_amount")) or Decimal("0")
        # 待核销为整期滚动余额 → 合计取末期(最后一行=最后一期)余额
        pend = _num(row.get("pending_writeoff"))
        if pend is not None:
            last_pending = pend
        r += 1

    total_fill = PatternFill("solid", fgColor="FDE9D9")
    total_vals = ["合计", "", "", float(_q(s_hx)), float(_q(last_pending)), float(_q(s_jy)), float(_q(s_fee)), s_night, "", float(_q(s_repay))]
    for c, v in enumerate(total_vals, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(bold=True); cell.alignment = center; cell.fill = total_fill; cell.border = border

    for i, wd in enumerate([8, 30, 20, 15, 15, 15, 13, 8, 13, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = wd

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
