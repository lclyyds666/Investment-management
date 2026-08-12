"""审批单打印导出：服务端填充原始 xlsx 模板（格式高度还原）。

用 openpyxl 加载 backend/app/templates/approval/{payment,business}.xlsx 原件，
仅向数据单元格写值，保留模板全部合并单元格/边框/字体/行高。签批栏写审批人姓名
（及意见），若装有 Pillow 则叠加电子签名图片；无 Pillow 时降级为姓名文本，永不报错。
"""
from __future__ import annotations

import base64
import io
import logging
from pathlib import Path

import openpyxl

from app.core.enums import ContractType

logger = logging.getLogger("app.approval_print")

_TPL_DIR = Path(__file__).resolve().parent.parent / "templates" / "approval"

# 是否可嵌入签名图片（生产装 Pillow 即启用；缺失则降级姓名文本）
try:  # noqa: SIM105
    from openpyxl.drawing.image import Image as _XLImage  # noqa: F401
    from PIL import Image as _PILImage  # noqa: F401

    _CAN_EMBED_IMG = True
except Exception:  # noqa: BLE001
    _CAN_EMBED_IMG = False


# 数据单元格映射（值写到合并区左上角锚点）
_DATA_CELLS = {
    ContractType.PAYMENT: {
        "department": "B2", "apply_date": "F2", "customer_name": "C3",
        "business_type": "C4", "business_desc": "C5", "contract_no": "F5",
        "amount": "C6", "remark": "C7", "bank_name": "C8", "bank_account": "F8",
    },
    ContractType.BUSINESS: {
        "department": "B2", "apply_date": "F2", "customer_name": "C3",
        "business_type": "C4", "business_desc": "C5", "contract_no": "F5",
    },
}

# 签批栏按岗位代码映射。仅历史 v1 行在无岗位快照时回退旧 role。
_SIGN_CELLS = {
    ContractType.PAYMENT: {
        "supply.business_handler": "C9",
        "supply.business_reviewer": "F9",
        "supply.finance_handler": "C10",
        "supply.company_leader": "F10",
        "investment.duty.supply_risk_review": "C11",
        "investment.duty.supply_finance_review": "F11",
        "governance.supply_leader": "C12",
    },
    ContractType.BUSINESS: {
        "supply.business_handler": "C6",
        "supply.business_reviewer": "F6",
        "supply.company_leader": "C7",
        "investment.duty.supply_risk_review": "C8",
        "governance.supply_leader": "C9",
    },
}

_LEGACY_ROLE_POSITIONS = {
    "business_handler": "supply.business_handler",
    "business_reviewer": "supply.business_reviewer",
    "finance_handler": "supply.finance_handler",
    "scm_director": "supply.company_leader",
    "risk_auditor": "investment.duty.supply_risk_review",
    "finance_reviewer": "investment.duty.supply_finance_review",
    "invest_director": "governance.supply_leader",
}


def _template_path(form_type: ContractType) -> Path:
    return _TPL_DIR / ("payment.xlsx" if form_type == ContractType.PAYMENT else "business.xlsx")


def _embed_signature(ws, cell: str, data_uri: str) -> bool:
    """将 data-URI 签名图片嵌入指定单元格；成功返回 True。"""
    if not (_CAN_EMBED_IMG and data_uri and data_uri.startswith("data:")):
        return False
    try:
        b64 = data_uri.split(",", 1)[1]
        raw = base64.b64decode(b64)
        img = _XLImage(io.BytesIO(raw))
        img.width, img.height = 110, 44  # 控制签章尺寸，避免撑大行高
        ws.add_image(img, cell)
        return True
    except Exception as exc:  # noqa: BLE001
        logger.warning("签名图片嵌入失败，降级文本：%s", exc)
        return False


def build_approval_form_xlsx(form: dict, steps: list[dict]) -> bytes:
    """填充审批单模板并返回 xlsx 字节。

    form: 审批单字段字典（含 form_type: ContractType）。
    steps: 每级审批动作 [{step, name, comment, signature, date, action}]。
    """
    form_type: ContractType = form["form_type"]
    wb = openpyxl.load_workbook(_template_path(form_type))
    ws = wb["Sheet1"]
    ws["B12" if form_type == ContractType.PAYMENT else "B9"] = "供管公司分管领导"

    cells = _DATA_CELLS[form_type]

    def _set(key: str, value):
        cell = cells.get(key)
        if cell is not None:
            ws[cell] = value

    _set("department", form.get("department") or "供管公司")
    _set("apply_date", str(form.get("apply_date") or ""))
    _set("customer_name", form.get("customer_name") or "")
    _set("business_type", form.get("business_type") or "")
    _set("business_desc", form.get("business_desc") or "详见合同")
    _set("contract_no", form.get("contract_no") or "")

    if form_type == ContractType.PAYMENT:
        amount = form.get("amount")
        words = form.get("amount_words") or ""
        amount_txt = f"小写：¥{amount}    大写：{words}" if amount is not None else ""
        _set("amount", amount_txt)
        remark = (form.get("remark") or "").strip()
        _set("remark", f"备注：{remark}" if remark else "备注：")
        _set("bank_name", form.get("bank_name") or "")
        _set("bank_account", form.get("bank_account") or "")

    # 新流程按岗位快照定位签批栏；历史 v1 行才读取旧 role。
    by_position = {}
    for item in steps:
        position_code = item.get("position_code")
        if not position_code:
            position_code = _LEGACY_ROLE_POSITIONS.get(item.get("role") or "")
        if position_code:
            by_position[position_code] = item
    sign_cells = _SIGN_CELLS[form_type]
    for position_code, cell in sign_cells.items():
        act = by_position.get(position_code)
        if not act:
            continue
        name = act.get("name") or ""
        comment = (act.get("comment") or "").strip()
        embedded = _embed_signature(ws, cell, act.get("signature") or "")
        # 文本：优先意见+姓名；已嵌图则仅留意见（图即签名），避免覆盖
        parts = []
        if comment:
            parts.append(comment)
        if not embedded and name:
            parts.append(name)
        text = "  ".join(parts)
        if text:
            ws[cell] = text

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
