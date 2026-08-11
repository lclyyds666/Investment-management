"""文旅业务·门票平台核销台账 Excel 适配器。

职责：
1. 解析平台对账明细 xlsx（同一文件可包含抖音/美团/携程/同程多个 Sheet），按景区与
   平台的固定策略计算服务商到账，并解析对账周期跨度 / 核对日期文本。
2. 使用调用方传入的景区配置快照计算服务商佣金、景区核销、结算金额和服务费，
   另按期次递推出景区待核销金额(滚动余额，见 running_pending)。
3. 用 openpyxl 生成标准格式业务台账 xlsx（含合计行）供导出。

列一律按「表头名」定位，兼容明细表 70/72 列的差异；金额解析宽松（去 ¥、逗号等）。
"""
from __future__ import annotations

import json
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO

import openpyxl

from app.services.ledger_calculator import (
    DEFAULT_COMMISSION_RATE as CALC_DEFAULT_COMMISSION_RATE,
    DEFAULT_RATE_HEXIAO as CALC_DEFAULT_RATE_HEXIAO,
    DEFAULT_RATE_SETTLE as CALC_DEFAULT_RATE_SETTLE,
    calculate_running_balances,
    calculate_ticket_ledger as _calculate_ticket_ledger,
    quantize_money as _q,
    running_pending as _running_pending,
)

# 门票产品仅作为缺失产品名称时的系统兜底；实际景区由 scenic_id 作用域决定。
DEFAULT_TICKET_PRODUCT = "水上世界/童话世界/海洋王国"
DEFAULT_RATE_HEXIAO = CALC_DEFAULT_RATE_HEXIAO  # 景区核销率
# 结算费率：结算金额 = 出版应得B × 结算费率。默认 0.94（= 旧核销率0.90 + 旧服务费率0.04），
# 保证历史/现有台账数值完全不变；服务费改为派生 = 结算金额 − 景区核销金额。
DEFAULT_RATE_SETTLE = CALC_DEFAULT_RATE_SETTLE
DEFAULT_RATE_FEE = Decimal("0.04")     # 旧服务费率（保留常量，仅历史/迁移回填参考）
# 服务商佣金默认率(对订单实收金额)：服务商佣金 = 订单实收×6% − 达人服务费 − 团长服务费
DEFAULT_COMMISSION_RATE = CALC_DEFAULT_COMMISSION_RATE

# 明细表关键列的表头名（按名定位，抗列数差异）
COL_SHISHOU = "订单实收金额"
COL_RUANJIAN = "软件服务费"
COL_DAREN = "达人服务费"
COL_TUANZHANG = "团长服务费"
COL_FUWUSHANG = "服务商服务费"
COL_HEXIAO_TIME = "核销时间"
# 携程门票对账明细。携程没有抖音的订单实收/佣金列，结算价即服务商到账口径。
COL_XC_JIESUAN = "结算价金额"
COL_XC_FLOW_TYPE = "流水类型"
COL_XC_SERVICE_DATE = "服务完成日期"
COL_XC_DEPARTURE_DATE = "出发时间"
COL_XC_PAYMENT_DATE = "付款日期"
XC_ORDER_COST = "订单成本"
# 美团门票消费结算明细。应付金额为扣除技术服务费后的商家结算净额。
COL_MT_AMOUNT = "应付金额"
COL_MT_TECH_FEE = "技术服务费"
COL_MT_SETTLE_TYPE = "结算方式"
COL_MT_COUNT = "张数"
COL_MT_TIME = "时间"
MT_CONSUMPTION_SETTLEMENT = "消费结算"
# 同程门票结算明细。商家应收为平台最终结算给商家的净额。
COL_TC_AMOUNT = "商家应收"
COL_TC_COUNT = "订单票数"
COL_TC_DATE = "旅游日期"
_HEADER_SCAN_ROWS = 20
# 服务商到账金额 = 订单实收 − 软件 − 达人 − 团长（明细中费用列为负数，直接相加）。
# 注意：明细里的「服务商服务费」列其实是 -(服务商到账金额)，若一并相加会把结果抵消为 0，故不纳入。
_FEE_COLS = (COL_RUANJIAN, COL_DAREN, COL_TUANZHANG)

# 公式不属于运营配置，只能随代码评审和发布调整。
_RECEIVED_RULES = {
    ("zunyi-zoo", "抖音"): "zunyi_douyin",
    ("zunyi-zoo", "美团"): "zunyi_meituan",
    ("nanyang-wildlife", "抖音"): "nanyang_douyin",
}


def _num(v):
    """宽松转 Decimal（去 ¥、逗号、%、空格等）；无法解析返回 None。"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            return Decimal(str(v))
        except InvalidOperation:
            return None
    s = re.sub(r"[^\d.\-]", "", str(v).strip())
    if not s or s in ("-", ".", "-."):
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _to_date(v):
    """单元格转 date；解析失败返回 None。"""
    if v is None or str(v).strip() == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    m = re.search(r"(\d{4})\D+(\d{1,2})\D+(\d{1,2})", s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    return None


def _header_index(header: list, name: str) -> int:
    """在表头行里按名找列索引；找不到返回 -1。"""
    for i, h in enumerate(header):
        if h is not None and str(h).strip() == name:
            return i
    return -1


def _detect_platform(header: list) -> str | None:
    """根据关键表头识别平台，不依赖 Sheet 名称。"""
    names = {
        str(value).strip()
        for value in header
        if value is not None and str(value).strip()
    }
    signatures = (
        ("抖音", {COL_SHISHOU, COL_HEXIAO_TIME}),
        ("美团", {COL_MT_AMOUNT, COL_MT_SETTLE_TYPE, COL_MT_COUNT, COL_MT_TIME}),
        ("携程", {COL_XC_JIESUAN, COL_XC_FLOW_TYPE}),
        ("同程", {COL_TC_AMOUNT, COL_TC_COUNT, COL_TC_DATE}),
    )
    return next(
        (platform for platform, required in signatures if required.issubset(names)),
        None,
    )


def _find_platform_header(rows_iter) -> tuple[str, list] | None:
    """在 Sheet 前若干行查找可识别的平台表头，并保留后续明细迭代器位置。"""
    for _, raw in zip(range(_HEADER_SCAN_ROWS), rows_iter):
        if not raw or not any(value is not None and str(value).strip() for value in raw):
            continue
        header = list(raw)
        platform = _detect_platform(header)
        if platform:
            return platform, header
    return None


def _period_from_filename(
    filename: str,
    reference_start: date | None = None,
    reference_end: date | None = None,
) -> tuple[date | None, date | None]:
    """从文件名解析对账周期，如 对账明细-2026.04.29-2026.05.19.xlsx。"""
    if not filename:
        return None, None
    m = re.search(
        r"(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})\D+(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})",
        filename,
    )
    if m:
        try:
            d1 = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            d2 = date(int(m.group(4)), int(m.group(5)), int(m.group(6)))
            return d1, d2
        except ValueError:
            pass

    # 兼容「遵义动物园5.25-6.21.xlsx」这类省略年份的业务文件名；
    # 年份取明细日期，跨年时自动把结束日期放到下一年。
    reference = reference_start or reference_end
    if reference is None:
        return None, None
    short_match = re.search(
        r"(?<!\d)(\d{1,2})[.\-/](\d{1,2})\D+(\d{1,2})[.\-/](\d{1,2})(?!\d)",
        filename or "",
    )
    if not short_match:
        return None, None
    try:
        start_month, start_day, end_month, end_day = map(int, short_match.groups())
        start = date(reference.year, start_month, start_day)
        end_year = reference.year + int((end_month, end_day) < (start_month, start_day))
        end = date(end_year, end_month, end_day)
        return start, end
    except ValueError:
        return None, None


def _row_count(value, default: int = 1) -> int:
    """平台聚合行换算为核销票数；缺失时按一条核销记录处理。"""
    parsed = _num(value)
    if parsed is None:
        return default
    return max(int(parsed), 0)


def _fee_charge(value: Decimal | None) -> Decimal:
    """平台费用列正负号不统一，业务公式统一按正向费用金额扣减。"""
    return abs(value or Decimal("0"))


def parse_reconciliation(
    content: bytes,
    filename: str = "",
    *,
    scenic_id: str = "legacy",
    rate_hexiao: Decimal = DEFAULT_RATE_HEXIAO,
    rate_settle: Decimal = DEFAULT_RATE_SETTLE,
    commission_rate: Decimal = DEFAULT_COMMISSION_RATE,
    commission_override=None,
    ticket_product: str = DEFAULT_TICKET_PRODUCT,
) -> dict:
    """解析一个对账明细 xlsx，返回汇总。

    返回:
      {
        "supplier_received": Decimal,   # 服务商到账金额 = Σ(订单实收 - 软件 - 达人 - 团长)
        "order_count": int,             # 有效核销订单数
        "period_start": date|None,
        "period_end": date|None,
        "period_text": str,             # 对账周期文本，如 2026/4/29-2026/5/19
        "check_date_text": str,         # 核对日期（同 period_text，供台账「核对日期」列）
        "sheets": [sheet 名列表],
      }

    对账周期优先取自文件名（如 对账明细-2026.04.29-2026.05.19.xlsx）；
    文件名无法解析时，回退到核销时间跨度。
    """
    # read_only=True 流式读取，内存可控；用 try/finally 保证异常时也释放工作簿
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)

    # 同一文件可同时包含多个平台；每个平台独立生成一条台账草稿。
    aggregates: dict[str, dict] = {}

    def platform_aggregate(platform: str) -> dict:
        return aggregates.setdefault(platform, {
            "supplier_received": Decimal("0"),
            "order_count": 0,
            "positive_count": 0,
            "min_dt": None,
            "max_dt": None,
            "sheets": [],
            "daily": {},
        })

    def add_date(aggregate: dict, value: date | None) -> None:
        if value is None:
            return
        aggregate["min_dt"] = (
            value if aggregate["min_dt"] is None else min(aggregate["min_dt"], value)
        )
        aggregate["max_dt"] = (
            value if aggregate["max_dt"] is None else max(aggregate["max_dt"], value)
        )

    try:
        for ws in wb.worksheets:
            rows_iter = ws.iter_rows(values_only=True)
            detected = _find_platform_header(rows_iter)
            if detected is None:
                continue
            platform, header = detected

            if platform == "抖音":
                aggregate = platform_aggregate("抖音")
                aggregate["sheets"].append(ws.title)
                i_shishou = _header_index(header, COL_SHISHOU)
                i_fees = [_header_index(header, c) for c in _FEE_COLS]
                i_fuwushang = _header_index(header, COL_FUWUSHANG)
                i_time = _header_index(header, COL_HEXIAO_TIME)
                received_rule = _RECEIVED_RULES.get((scenic_id, "抖音"), "default")
                if received_rule == "zunyi_douyin":
                    missing = [
                        name for name, idx in (
                            (COL_DAREN, i_fees[1]), (COL_FUWUSHANG, i_fuwushang)
                        ) if idx < 0
                    ]
                    if missing:
                        raise ValueError(
                            f"遵义动物园抖音明细缺少必要列：{'、'.join(missing)}"
                        )
                for raw in rows_iter:
                    if not raw:
                        continue
                    shishou = _num(raw[i_shishou]) if i_shishou < len(raw) else None
                    fee_vals = [
                        (_num(raw[idx]) if 0 <= idx < len(raw) else None) for idx in i_fees
                    ]
                    fuwushang = (
                        _num(raw[i_fuwushang]) if 0 <= i_fuwushang < len(raw) else None
                    )
                    # 空行 / 小计行：实收与全部费用都无值 → 跳过
                    if shishou is None and all(f is None for f in fee_vals) and fuwushang is None:
                        continue
                    if received_rule == "zunyi_douyin":
                        base = (
                            (shishou or Decimal("0"))
                            - _fee_charge(fee_vals[1])
                            - _fee_charge(fuwushang)
                        )
                    elif received_rule == "nanyang_douyin":
                        base = shishou or Decimal("0")
                    else:
                        base = shishou or Decimal("0")
                        for fee in fee_vals:
                            base += fee or Decimal("0")  # 通用账单费用为负数，直接相加
                    aggregate["supplier_received"] += base
                    aggregate["order_count"] += 1
                    if shishou is not None and shishou > 0:
                        aggregate["positive_count"] += 1

                    d = _to_date(raw[i_time]) if 0 <= i_time < len(raw) else None
                    add_date(aggregate, d)
                    key = d.isoformat() if d else "NA"
                    dd = aggregate["daily"].setdefault(key, {
                        "received": Decimal("0"), "shishou": Decimal("0"),
                        "daren": Decimal("0"), "tuanzhang": Decimal("0"),
                    })
                    dd["received"] += base
                    dd["shishou"] += (shishou or Decimal("0"))
                    dd["daren"] += (fee_vals[1] or Decimal("0"))
                    dd["tuanzhang"] += (fee_vals[2] or Decimal("0"))
                continue

            if platform == "携程":
                aggregate = platform_aggregate("携程")
                aggregate["sheets"].append(ws.title)
                i_xc_jiesuan = _header_index(header, COL_XC_JIESUAN)
                i_flow = _header_index(header, COL_XC_FLOW_TYPE)
                i_count = _header_index(header, "使用份数")
                i_service = _header_index(header, COL_XC_SERVICE_DATE)
                i_departure = _header_index(header, COL_XC_DEPARTURE_DATE)
                i_payment = _header_index(header, COL_XC_PAYMENT_DATE)
                for raw in rows_iter:
                    if not raw:
                        continue
                    # 携程账单可能混有调账等流水，只将订单成本计入核销台账。
                    if 0 <= i_flow < len(raw):
                        flow_type = str(raw[i_flow] or "").strip()
                        if flow_type and flow_type != XC_ORDER_COST:
                            continue
                    base = _num(raw[i_xc_jiesuan]) if i_xc_jiesuan < len(raw) else None
                    if base is None:
                        continue
                    count = _row_count(raw[i_count] if 0 <= i_count < len(raw) else None)
                    d = None
                    for idx in (i_service, i_departure, i_payment):
                        if 0 <= idx < len(raw):
                            d = _to_date(raw[idx])
                            if d is not None:
                                break
                    aggregate["supplier_received"] += base
                    aggregate["order_count"] += count
                    if base > 0:
                        aggregate["positive_count"] += count
                    add_date(aggregate, d)
                    key = d.isoformat() if d else "NA"
                    dd = aggregate["daily"].setdefault(key, {
                        "received": Decimal("0"), "shishou": Decimal("0"),
                        "daren": Decimal("0"), "tuanzhang": Decimal("0"),
                    })
                    dd["received"] += base
                continue

            if platform == "美团":
                aggregate = platform_aggregate("美团")
                aggregate["sheets"].append(ws.title)
                i_amount = _header_index(header, COL_MT_AMOUNT)
                i_tech_fee = _header_index(header, COL_MT_TECH_FEE)
                i_settle_type = _header_index(header, COL_MT_SETTLE_TYPE)
                i_count = _header_index(header, COL_MT_COUNT)
                i_time = _header_index(header, COL_MT_TIME)
                received_rule = _RECEIVED_RULES.get((scenic_id, "美团"), "default")
                if received_rule == "zunyi_meituan" and i_tech_fee < 0:
                    raise ValueError(f"遵义动物园美团明细缺少必要列：{COL_MT_TECH_FEE}")
                for raw in rows_iter:
                    if not raw:
                        continue
                    settle_type = (
                        str(raw[i_settle_type] or "").strip()
                        if 0 <= i_settle_type < len(raw)
                        else ""
                    )
                    if settle_type != MT_CONSUMPTION_SETTLEMENT:
                        continue
                    base = _num(raw[i_amount]) if 0 <= i_amount < len(raw) else None
                    if base is None:
                        continue
                    if received_rule == "zunyi_meituan":
                        tech_fee = (
                            _num(raw[i_tech_fee]) if 0 <= i_tech_fee < len(raw) else None
                        )
                        base += tech_fee or Decimal("0")
                    count = _row_count(raw[i_count] if 0 <= i_count < len(raw) else None)
                    d = _to_date(raw[i_time]) if 0 <= i_time < len(raw) else None
                    add_date(aggregate, d)
                    aggregate["supplier_received"] += base
                    aggregate["order_count"] += count
                    if base > 0:
                        aggregate["positive_count"] += count
                    key = d.isoformat() if d else "NA"
                    dd = aggregate["daily"].setdefault(key, {
                        "received": Decimal("0"), "shishou": Decimal("0"),
                        "daren": Decimal("0"), "tuanzhang": Decimal("0"),
                    })
                    dd["received"] += base
                continue

            if platform == "同程":
                aggregate = platform_aggregate("同程")
                aggregate["sheets"].append(ws.title)
                i_amount = _header_index(header, COL_TC_AMOUNT)
                i_count = _header_index(header, COL_TC_COUNT)
                i_date = _header_index(header, COL_TC_DATE)
                for raw in rows_iter:
                    if not raw:
                        continue
                    base = _num(raw[i_amount]) if 0 <= i_amount < len(raw) else None
                    if base is None:
                        continue
                    count = _row_count(raw[i_count] if 0 <= i_count < len(raw) else None)
                    d = _to_date(raw[i_date]) if 0 <= i_date < len(raw) else None
                    add_date(aggregate, d)
                    aggregate["supplier_received"] += base
                    aggregate["order_count"] += count
                    if base > 0:
                        aggregate["positive_count"] += count
                    key = d.isoformat() if d else "NA"
                    dd = aggregate["daily"].setdefault(key, {
                        "received": Decimal("0"), "shishou": Decimal("0"),
                        "daren": Decimal("0"), "tuanzhang": Decimal("0"),
                    })
                    dd["received"] += base
    finally:
        wb.close()

    # 周期优先取文件名；否则各平台分别回退到有效业务日期跨度。
    platforms = []
    for platform in ("抖音", "美团", "携程", "同程"):
        if platform not in aggregates:
            continue
        aggregate = aggregates[platform]
        fn_start, fn_end = _period_from_filename(
            filename, aggregate["min_dt"], aggregate["max_dt"]
        )
        p_start = fn_start or aggregate["min_dt"]
        p_end = fn_end or aggregate["max_dt"]
        period_text = ""
        if p_start and p_end:
            period_text = (
                f"{p_start.year}/{p_start.month}/{p_start.day}-"
                f"{p_end.year}/{p_end.month}/{p_end.day}"
            )
        defs = daily_defaults(
            aggregate["daily"],
            rate_hexiao=rate_hexiao,
            rate_settle=rate_settle,
            commission_override=commission_override,
            commission_rate=commission_rate,
            platform=platform,
            scenic_id=scenic_id,
        )
        platforms.append({
            "platform": platform,
            "ticket_product": ticket_product,
            "rate_hexiao": rate_hexiao,
            "rate_settle": rate_settle,
            "commission_rate": commission_rate,
            "supplier_received": _q(aggregate["supplier_received"]),
            "suggested_commission": defs["commission"],
            "def_hexiao": defs["hexiao"],
            "def_service_fee": defs["service_fee"],
            "def_jinying": defs["jinying"],
            "daily_json": serialize_daily(aggregate["daily"]),
            "order_count": aggregate["order_count"],
            "positive_count": aggregate["positive_count"],
            "period_start": p_start,
            "period_end": p_end,
            "period_text": period_text,
            "check_date_text": period_text,
            "sheets": aggregate["sheets"],
        })

    # 保留旧的顶层返回字段，避免只含抖音的既有调用和历史逐日明细恢复受影响。
    legacy = next((item for item in platforms if item["platform"] == "抖音"), None)
    legacy = legacy or (platforms[0] if platforms else {
        "platform": "", "supplier_received": Decimal("0"),
        "suggested_commission": Decimal("0"), "def_hexiao": Decimal("0"),
        "def_service_fee": Decimal("0"), "def_jinying": Decimal("0"),
        "daily_json": "", "order_count": 0, "positive_count": 0,
        "period_start": None, "period_end": None, "period_text": "",
        "check_date_text": "", "sheets": [],
    })
    return {**legacy, "platforms": platforms}


# --------------------------------------------------------------------------- #
# 逐日明细：序列化持久化 + 逐日重算（编辑改费率/佣金时仍按天累加，不退回总额×费率）
# --------------------------------------------------------------------------- #
def _days_from_daily(daily: dict[str, dict]) -> list[dict]:
    """按日聚合 dict → 逐日列表（Decimal）。"""
    return [{
        "recv": dd["received"], "shishou": dd["shishou"],
        "daren": dd["daren"], "tuanzhang": dd["tuanzhang"],
    } for dd in daily.values()]


def serialize_daily(daily: dict[str, dict]) -> str:
    """逐日明细序列化为 JSON（Decimal→字符串），随台账行持久化，供编辑时逐日重算。"""
    out = [{
        "r": str(dd["received"]), "s": str(dd["shishou"]),
        "d": str(dd["daren"]), "t": str(dd["tuanzhang"]),
    } for dd in daily.values()]
    return json.dumps(out, ensure_ascii=False)


def _days_from_json(daily_json: str) -> list[dict]:
    if not daily_json:
        return []
    try:
        raw = json.loads(daily_json)
    except (ValueError, TypeError):
        return []
    days = []
    for d in raw:
        commission_inputs = {
            key: d[key]
            for key in (
                "commission_shishou", "commission_daren", "commission_tuanzhang",
                "cs", "cd", "ct",
            )
            if key in d
        }
        days.append({
            "recv": _num(d.get("r")) or Decimal("0"),
            "shishou": _num(d.get("s")) or Decimal("0"),
            "daren": _num(d.get("d")) or Decimal("0"),
            "tuanzhang": _num(d.get("t")) or Decimal("0"),
            **commission_inputs,
        })
    return days


def recompute_from_days(days: list[dict],
                        rate_hexiao: Decimal = DEFAULT_RATE_HEXIAO,
                        rate_settle: Decimal = DEFAULT_RATE_SETTLE,
                        commission_override=None,
                        commission_rate: Decimal = DEFAULT_COMMISSION_RATE,
                        platform: str = "抖音",
                        scenic_id: str = "legacy") -> dict | None:
    return _calculate_ticket_ledger(
        scenic_id, days, rate_hexiao=rate_hexiao, rate_settle=rate_settle,
        commission_override=commission_override, commission_rate=commission_rate,
        platform=platform,
    )


def recompute_from_json(daily_json: str,
                        rate_hexiao: Decimal = DEFAULT_RATE_HEXIAO,
                        rate_settle: Decimal = DEFAULT_RATE_SETTLE,
                        commission_override=None,
                        commission_rate: Decimal = DEFAULT_COMMISSION_RATE,
                        platform: str = "抖音",
                        scenic_id: str = "legacy") -> dict | None:
    return recompute_from_days(_days_from_json(daily_json), rate_hexiao, rate_settle,
                               commission_override, commission_rate, platform, scenic_id)


def calculate_ticket_ledger(scenic_id: str, excel_data, **kwargs) -> dict | None:
    """公共门票计算入口；Excel 解析结果由调用方传入，计算本身无副作用。"""
    return _calculate_ticket_ledger(scenic_id, excel_data, **kwargs)


calculateTicketLedger = calculate_ticket_ledger


def daily_defaults(daily: dict[str, dict],
                   rate_hexiao: Decimal = DEFAULT_RATE_HEXIAO,
                   rate_settle: Decimal = DEFAULT_RATE_SETTLE,
                   commission_override=None,
                   commission_rate: Decimal = DEFAULT_COMMISSION_RATE,
                   platform: str = "抖音",
                   scenic_id: str = "legacy") -> dict:
    """解析时的按日精准默认值（佣金取逐日自动值）。"""
    res = recompute_from_days(
        _days_from_daily(daily), rate_hexiao, rate_settle, commission_override,
        commission_rate, platform, scenic_id
    )
    if res is None:
        return {"commission": Decimal("0"), "hexiao": Decimal("0"),
                "service_fee": Decimal("0"), "jinying": Decimal("0")}
    return {"commission": res["supplier_commission"], "hexiao": res["hexiao_amount"],
            "service_fee": res["service_fee"], "jinying": res["jinying_amount"]}


def compute_row(
    supplier_received: Decimal,
    supplier_commission: Decimal = Decimal("0"),
    rate_hexiao: Decimal = DEFAULT_RATE_HEXIAO,
    rate_settle: Decimal = DEFAULT_RATE_SETTLE,
    platform: str = "抖音",
    scenic_id: str = "legacy",
) -> dict:
    """由服务商到账、服务商佣金与比例计算台账计算列（无逐日明细时的期级兜底）。

      出版应得到账金额 B = 服务商到账 - 服务商佣金
      景区核销金额 = B × 核销率
      结算金额     = B × 结算费率
      服务费       = 结算金额 − 景区核销金额
    服务商佣金仅对抖音生效（其它平台佣金恒 0）。
    """
    return _calculate_ticket_ledger(
        scenic_id, [], supplier_received=supplier_received,
        commission_override=supplier_commission, rate_hexiao=rate_hexiao,
        rate_settle=rate_settle, platform=platform,
    )


def running_pending(prev_balance: Decimal, payment_amount: Decimal, hexiao_amount: Decimal) -> Decimal:
    """期次递推：本期景区待核销金额 = 上期剩余余额 + 本期付款金额 - 本期景区核销金额。

    首期时 prev_balance 传 0 即为「首期付款金额 - 首期景区核销金额」。
    """
    return _running_pending(prev_balance, payment_amount, hexiao_amount)


# 导出台账列顺序（对齐手工业务台账样表；景区待核销金额紧邻景区核销金额右侧）
# 注：付款金额不在生成台账中展示（仅待确认台账录入 + 参与后端递推），故导出亦不含该列。
_EXPORT_HEADERS = [
    "平台", "景区门票", "核对日期",
    "景区核销金额", "景区待核销金额", "结算金额", "服务费",
    "回款日期", "回款金额",
]


def _fmt_date(v) -> str:
    if not v:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v)


def _fmt_amount(v):
    if v is None or v == "":
        return ""
    d = _num(v)
    return float(_q(d)) if d is not None else str(v)


def build_export_workbook(rows: list[dict], title: str = "业务台账") -> bytes:
    """生成标准格式业务台账 xlsx（含标题行 + 表头 + 数据 + 合计行）。

    rows 每项字段：pay_date, platform, ticket_product, check_date_text,
    hexiao_amount, jinying_amount, service_fee, repay_date, repay_amount。
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "业务台账"

    ncol = len(_EXPORT_HEADERS)
    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 标题行（合并）
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    tc = ws.cell(row=1, column=1, value=title)
    tc.font = Font(size=13, bold=True)
    tc.alignment = center

    # 表头
    head_fill = PatternFill("solid", fgColor="DCE6F1")
    for c, name in enumerate(_EXPORT_HEADERS, start=1):
        cell = ws.cell(row=2, column=c, value=name)
        cell.font = Font(bold=True)
        cell.alignment = center
        cell.fill = head_fill
        cell.border = border

    # 数据行
    sum_hexiao = Decimal("0")
    sum_pending = Decimal("0")
    sum_jinying = Decimal("0")
    sum_fee = Decimal("0")
    sum_repay = Decimal("0")
    r = 3
    for row in rows:
        vals = [
            row.get("platform", "") or "",
            row.get("ticket_product", "") or DEFAULT_TICKET_PRODUCT,
            row.get("check_date_text", "") or "",
            _fmt_amount(row.get("hexiao_amount")),
            _fmt_amount(row.get("pending_writeoff")),
            _fmt_amount(row.get("jinying_amount")),
            _fmt_amount(row.get("service_fee")),
            _fmt_date(row.get("repay_date")),
            _fmt_amount(row.get("repay_amount")),
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = center
            cell.border = border
        sum_hexiao += _num(row.get("hexiao_amount")) or Decimal("0")
        sum_jinying += _num(row.get("jinying_amount")) or Decimal("0")
        sum_fee += _num(row.get("service_fee")) or Decimal("0")
        sum_repay += _num(row.get("repay_amount")) or Decimal("0")
        r += 1
    # 景区待核销金额为滚动余额，合计取末期余额（最后一行的值）
    if rows:
        sum_pending = _num(rows[-1].get("pending_writeoff")) or Decimal("0")

    # 合计行
    total_fill = PatternFill("solid", fgColor="FDE9D9")
    total_vals = [
        "合计", "", "",
        float(_q(sum_hexiao)), float(_q(sum_pending)),
        float(_q(sum_jinying)), float(_q(sum_fee)),
        "", float(_q(sum_repay)),
    ]
    for c, v in enumerate(total_vals, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(bold=True)
        cell.alignment = center
        cell.fill = total_fill
        cell.border = border

    # 列宽
    widths = [8, 22, 20, 15, 15, 15, 14, 13, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
