"""供管公司项目统计表（Sheet2）预处理与入库。

职责：
1. 加载年度项目投入及回款收益统计表的 Sheet2，解析每个项目的
   投入金额 / 回款小计 / 实现毛利 / 收益率 / 平台 / 付款日期 / 合同期限。
2. 原表单位为「万元」，入库统一 ×10000 转为元。
3. 按 (项目, 付款日期) UPSERT 写入 biz_project_metrics（覆盖式、幂等）。
"""
from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.project import ProjectMetrics

WAN = Decimal("10000")  # 万元 → 元

# Sheet2 列索引（两行合并表头，回款金额跨 1-4 季度 + 小计）
COL_SEQ = 0
COL_PROJECT = 1
COL_PLATFORM = 2
COL_INVEST = 3
COL_PAYDATE = 4
COL_TERM = 5
COL_HK_SUBTOTAL = 10   # 回款金额-小计
COL_PROFIT = 11        # 实现毛利
COL_RATE = 12          # 收益率


def _dec(v):
    """转 Decimal（去除逗号/空格等）；空或非数返回 None。"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            return Decimal(str(v))
        except InvalidOperation:
            return None
    s = re.sub(r"[^\d.\-]", "", str(v).strip())
    if not s or s in ("-", "."):
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _to_date(v):
    if v is None or str(v).strip() == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    m = re.search(r"(\d{4})\D+(\d{1,2})\D+(\d{1,2})", s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    return None


def parse_sheet2(content: bytes) -> list[dict]:
    """解析 Sheet2，返回项目指标列表（金额已转元）。"""
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb["Sheet2"] if "Sheet2" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    results = []
    for r in rows:
        if not r or len(r) <= COL_RATE:
            continue
        seq = r[COL_SEQ]
        # 仅接受序号为数字的项目数据行，跳过标题/表头/合计
        seq_num = _dec(seq)
        if seq_num is None:
            continue
        name = r[COL_PROJECT]
        if name is None or str(name).strip() == "":
            continue

        invest = _dec(r[COL_INVEST]) or Decimal("0")
        hk = _dec(r[COL_HK_SUBTOTAL])
        profit = _dec(r[COL_PROFIT])
        rate = _dec(r[COL_RATE])

        results.append({
            "seq": int(seq_num),
            "project_name": str(name).strip(),
            "platforms": str(r[COL_PLATFORM] or "").strip(),
            "invested_amount": invest * WAN,
            "realized_scale": (hk * WAN) if hk is not None else Decimal("0"),
            "gross_profit": (profit * WAN) if profit is not None else Decimal("0"),
            "profit_rate": rate,  # 保留小数原值
            "pay_date": _to_date(r[COL_PAYDATE]),
            "term_months": str(r[COL_TERM] or "").strip(),
        })
    return results


def upsert_projects(db: Session, parsed: list[dict]) -> list[ProjectMetrics]:
    """按 (项目, 付款日期) UPSERT，覆盖式、幂等。"""
    saved = []
    for item in parsed:
        row = db.scalar(
            select(ProjectMetrics).where(
                ProjectMetrics.project_name == item["project_name"],
                ProjectMetrics.pay_date == item["pay_date"],
            )
        )
        if not row:
            row = ProjectMetrics(
                project_name=item["project_name"], pay_date=item["pay_date"]
            )
            db.add(row)
        row.seq = item["seq"]
        row.platforms = item["platforms"]
        row.invested_amount = item["invested_amount"]
        row.realized_scale = item["realized_scale"]
        row.gross_profit = item["gross_profit"]
        row.profit_rate = item["profit_rate"]
        row.term_months = item["term_months"]
        saved.append(row)
    db.commit()
    for r in saved:
        db.refresh(r)
    return saved
